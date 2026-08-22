"""
Listing Factory: Multi-Marketplace Packaging Studio (v2.1)
===========================================================
A category-profile-driven local web application for e-commerce cataloging agencies.
Takes AI-generated JSON copy + loose image files, validates against Amazon.in,
Flipkart, and Meesho constraints across 13 distinct product families, organizes
images into SKU subfolders, populates marketplace mapping Excel workbooks, and outputs
a ready-to-deliver client ZIP archive with full cryptographic package audit metadata.

Version: Listing Factory v2.1 (Category-Profile-Driven)
Run:     python app.py
Open:    http://127.0.0.1:8000
"""

from __future__ import annotations

import glob
import hashlib
import io
import json
import os
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, List, Literal, Optional, Union

import pandas as pd
from fastapi import FastAPI, File, Form, UploadFile, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, field_validator
import uvicorn


# ──────────────────────────────────────────────
# Global Constants & Versioning
# ──────────────────────────────────────────────

TOOL_VERSION = "Listing Factory v2.1"
JSON_PROMPT_VERSION = "JSON Prompt v2.1 – Category Profile Architecture"
EXPECTED_SCHEMA_VERSION = "v2.0"
TO_BE_CONFIRMED = "To be confirmed"

STRUCTURAL_READINESS_DISCLAIMER = (
    "Structural completeness confirms schema and package checks only. It does not verify product facts, "
    "image content, tax classification, marketplace-policy compliance, or marketplace acceptance. "
    "Seller review is required before upload."
)

IMAGE_ROLE_DISCLAIMER = (
    "Image roles are assigned from filenames only. Listing Factory does not visually verify image content. "
    "The seller must confirm that each declared slot contains the intended image before marketplace upload."
)

MAX_SKUS_SOFT_LIMIT = 50
MAX_TOTAL_IMAGE_SIZE_MB = 200

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)


# ──────────────────────────────────────────────
# Category Profile Registry (13 Product Families)
# ──────────────────────────────────────────────

EXCLUDED_PROFILES = {
    "blouse": "Blouses",
    "lingerie": "Lingerie & Innerwear",
    "innerwear": "Innerwear",
    "shapewear": "Shapewear",
    "bra": "Bras & Intimates",
    "underwear": "Underwear & Briefs",
    "intimate_apparel": "Intimate Apparel",
}

CATEGORY_PROFILES: dict[str, dict[str, Any]] = {
    "women_ethnic_kurta": {
        "profile_id": "women_ethnic_kurta",
        "display_name": "Women Ethnic Wear (Kurta / Kurti / Tunic / Set)",
        "category_group": "Women Ethnic Wear",
        "amazon_item_type_keyword": "kurtas-and-ethnic-tops",
        "flipkart_vertical": "Ethnic Wear / Kurta",
        "meesho_category_hint": "Women Ethnic Wear / Kurtis",
        "target_gender_options": ["Women", "Girls"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["fabric", "sleeve", "neckline", "length", "pattern", "care_label", "product_type", "color", "sizes"],
        "optional_verified_fields": ["kurta_type", "occasion", "pocket", "closure", "waistline", "embroidery"],
        "controlled_attributes": {
            "fabric": ["Pure Cotton", "Rayon", "Georgette", "Silk Blend", "Crepe", "Chanderi Cotton", "Poly Cotton", "To be confirmed"],
            "kurta_type": ["Anarkali", "Straight", "A-line", "Flared", "Kaftan", "Frontslit", "Pathani", "To be confirmed"],
            "neck": ["Mandarin Neck", "Round Neck", "V-Neck", "Boat Neck", "Sweetheart Neck", "Collar Neck", "To be confirmed"],
            "sleeve": ["3/4 Sleeve", "Full Sleeve", "Half Sleeve", "Sleeveless", "Short Sleeve", "To be confirmed"],
            "length_type": ["Calf Length", "Ankle Length", "Knee Length", "Above Knee", "To be confirmed"],
            "pattern": ["Floral Print", "Solid", "Printed", "Embroidered", "Geometric Print", "Self Design", "Bandhani", "To be confirmed"],
            "occasion": ["Casual", "Festive", "Casual & Festive", "Party", "Formal", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Women's [Fabric] [Pattern] [Product Type] with [Verified Detail] ([Color])",
        "bullet_heading_guidance": ["100% PURE [FABRIC]", "[SILHOUETTE/UTILITY]", "COLORFAST & DURABLE", "VERSATILE STYLING", "PRECISE SIZING & CARE"],
        "supported": True,
        "excluded": False
    },
    "saree": {
        "profile_id": "saree",
        "display_name": "Sarees (Traditional / Daily / Festive)",
        "category_group": "Women Ethnic Wear",
        "amazon_item_type_keyword": "sarees",
        "flipkart_vertical": "Saree",
        "meesho_category_hint": "Women Ethnic Wear / Sarees",
        "target_gender_options": ["Women"],
        "age_group_options": ["Adults"],
        "required_verified_fields": ["fabric", "pattern", "saree_length", "blouse_piece_included", "care_label", "color", "occasion"],
        "optional_verified_fields": ["blouse_piece_fabric", "border_type", "pallu_type", "weave_or_technique", "embellishment", "package_contents"],
        "controlled_attributes": {
            "fabric": ["Georgette", "Chiffon", "Silk Blend", "Cotton", "Art Silk", "Crepe", "Chanderi", "Organza", "Poly Cotton", "To be confirmed"],
            "pattern": ["Floral Print", "Printed", "Solid", "Embroidered", "Woven Design", "Zari Work", "Bandhani", "Self Design", "To be confirmed"],
            "saree_type": ["Regular Saree", "Daily Wear Saree", "Festive Saree", "Party Wear Saree", "Traditional Saree", "To be confirmed"],
            "occasion": ["Casual", "Festive", "Party", "Wedding", "Formal", "To be confirmed"],
            "blouse_piece": ["Unstitched Blouse Piece", "Without Blouse Piece", "Stitched Blouse", "To be confirmed"],
            "saree_length": ["5.5 m", "6.0 m", "5.2 m", "6.3 m", "To be confirmed"],
            "border_type": ["Zari Border", "Embroidered Border", "Lace Border", "Printed Border", "Solid Border", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Women's [Fabric] [Pattern] [Saree Type] with [Blouse Piece Info] ([Color], [Saree Length])",
        "bullet_heading_guidance": ["[FABRIC] WEAVE", "PALLU & BORDER DESIGN", "COLORFAST & DURABLE", "DRAPING & OCCASIONS", "DIMENSIONS & CARE"],
        "supported": True,
        "excluded": False
    },
    "coord_set": {
        "profile_id": "coord_set",
        "display_name": "Co-ord Sets (Matching 2-Piece / 3-Piece Sets)",
        "category_group": "Women Western & Ethnic Sets",
        "amazon_item_type_keyword": "co-ord-sets",
        "flipkart_vertical": "Co-ords",
        "meesho_category_hint": "Women Western / Co-ord Sets",
        "target_gender_options": ["Women", "Girls"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["fabric", "pattern", "top_type", "bottom_type", "sleeve", "neckline", "top_length", "bottom_length", "care_label", "package_contents", "sizes", "color"],
        "optional_verified_fields": ["waist_rise", "closure_type", "pockets", "fit_type", "occasion"],
        "controlled_attributes": {
            "fabric": ["Pure Cotton", "Rayon", "Polyester Blend", "Crepe", "Linen Blend", "Satin", "Viscose", "To be confirmed"],
            "top_type": ["Tunic Top", "Crop Top", "Shirt Top", "Blazer Top", "Short Kurti", "To be confirmed"],
            "bottom_type": ["Palazzo", "Trousers", "Pants", "Shorts", "Skirts", "Joggers", "To be confirmed"],
            "sleeve": ["Short Sleeve", "3/4 Sleeve", "Full Sleeve", "Sleeveless", "To be confirmed"],
            "neckline": ["Collar Neck", "Round Neck", "V-Neck", "Mandarin Neck", "Square Neck", "To be confirmed"],
            "pattern": ["Solid", "Floral Print", "Abstract Print", "Striped", "Geometric Print", "To be confirmed"],
            "occasion": ["Casual", "Vacation / Resort", "Office Wear", "Party", "Festive", "To be confirmed"],
            "package_contents": ["1 Top, 1 Bottom", "1 Blazer, 1 Trouser", "1 Kurti, 1 Palazzo", "2 Piece Set", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Women's [Fabric] [Pattern] 2-Piece Co-ord Set with [Top Type] and [Bottom Type] ([Color])",
        "bullet_heading_guidance": ["COORDINATED [FABRIC] SET", "TOP & BOTTOM SPECIFICATIONS", "COLORFAST & DURABLE", "VERSATILE OCCASION WEAR", "SIZE GUIDE & CARE"],
        "supported": True,
        "excluded": False
    },
    "women_dress": {
        "profile_id": "women_dress",
        "display_name": "Women's Dresses (A-Line / Maxi / Midi / Fit & Flare)",
        "category_group": "Women Western Wear",
        "amazon_item_type_keyword": "dresses",
        "flipkart_vertical": "Dress",
        "meesho_category_hint": "Women Western / Dresses",
        "target_gender_options": ["Women", "Girls"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["fabric", "dress_type", "neckline", "sleeve", "length", "pattern", "care_label", "sizes", "color"],
        "optional_verified_fields": ["waistline", "hemline", "closure_type", "pockets", "layering", "fit_type", "occasion"],
        "controlled_attributes": {
            "dress_type": ["A-Line Dress", "Maxi Dress", "Midi Dress", "Fit and Flare Dress", "Shirt Dress", "Tiered Dress", "Bodycon", "To be confirmed"],
            "fabric": ["Pure Cotton", "Rayon", "Georgette", "Polyester Blend", "Crepe", "Linen Blend", "Viscose", "To be confirmed"],
            "neckline": ["Round Neck", "V-Neck", "Square Neck", "Collar Neck", "Sweetheart Neck", "Boat Neck", "To be confirmed"],
            "sleeve": ["Short Sleeve", "Sleeveless", "3/4 Sleeve", "Full Sleeve", "Puff Sleeve", "To be confirmed"],
            "length": ["Knee Length", "Calf Length", "Maxi / Ankle Length", "Mini Length", "To be confirmed"],
            "pattern": ["Floral Print", "Solid", "Printed", "Geometric Print", "Striped", "Polka Dot", "To be confirmed"],
            "occasion": ["Casual", "Party", "Vacation", "Formal / Work", "Festive", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Women's [Fabric] [Pattern] [Dress Type] with [Verified Detail] ([Color])",
        "bullet_heading_guidance": ["[FABRIC] CONSTRUCTION", "[DRESS TYPE] SILHOUETTE", "COLORFAST & DURABLE", "STYLING INSPIRATION", "FIT DIMENSIONS & CARE"],
        "supported": True,
        "excluded": False
    },
    "women_top": {
        "profile_id": "women_top",
        "display_name": "Women's Western Tops & Tunics",
        "category_group": "Women Western Wear",
        "amazon_item_type_keyword": "tunics-and-western-tops",
        "flipkart_vertical": "Top",
        "meesho_category_hint": "Women Western / Tops",
        "target_gender_options": ["Women", "Girls"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["fabric", "top_type", "neckline", "sleeve", "length", "pattern", "care_label", "sizes", "color"],
        "optional_verified_fields": ["closure_type", "pockets", "hemline", "fit_type", "occasion"],
        "controlled_attributes": {
            "top_type": ["Regular Top", "Peplum Top", "Crop Top", "Shirt Top", "Tunic", "Wrap Top", "To be confirmed"],
            "fabric": ["Pure Cotton", "Rayon", "Georgette", "Crepe", "Polyester Blend", "Viscose", "To be confirmed"],
            "neckline": ["Round Neck", "V-Neck", "Collar Neck", "Mandarin Neck", "Square Neck", "Boat Neck", "To be confirmed"],
            "sleeve": ["3/4 Sleeve", "Short Sleeve", "Full Sleeve", "Sleeveless", "Puff Sleeve", "To be confirmed"],
            "length": ["Regular Length", "Hip Length", "Crop Length", "Longline Length", "To be confirmed"],
            "pattern": ["Solid", "Floral Print", "Printed", "Striped", "Polka Dot", "Geometric Print", "To be confirmed"],
            "occasion": ["Casual", "Office Wear", "Party", "Casual & Festive", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Women's [Fabric] [Pattern] [Top Type] with [Verified Detail] ([Color])",
        "bullet_heading_guidance": ["PREMIUM [FABRIC] WEAVE", "[TOP TYPE] DESIGN DETAILS", "COLORFAST & DURABLE", "PAIRING SUGGESTIONS", "MEASUREMENTS & WASH CARE"],
        "supported": True,
        "excluded": False
    },
    "men_shirt": {
        "profile_id": "men_shirt",
        "display_name": "Men's Casual & Formal Shirts",
        "category_group": "Men Western Wear",
        "amazon_item_type_keyword": "mens-casual-shirts",
        "flipkart_vertical": "Shirt",
        "meesho_category_hint": "Men Western / Shirts",
        "target_gender_options": ["Men"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["fabric", "shirt_type", "collar_type", "sleeve", "pattern", "care_label", "sizes", "color"],
        "optional_verified_fields": ["fit_type", "pocket_count", "closure_type", "cuff_type", "hemline", "package_contents", "occasion"],
        "controlled_attributes": {
            "fabric": ["Pure Cotton", "Cotton Blend", "Linen Blend", "Denim", "Poplin Cotton", "Oxford Cotton", "To be confirmed"],
            "shirt_type": ["Casual Shirt", "Formal Shirt", "Denim Shirt", "Resort Shirt", "To be confirmed"],
            "collar_type": ["Spread Collar", "Button Down Collar", "Mandarin Collar", "Cutaway Collar", "Cuban Collar", "To be confirmed"],
            "sleeve": ["Full Sleeve", "Half Sleeve", "Roll-Up Sleeve", "Short Sleeve", "To be confirmed"],
            "pattern": ["Solid", "Checked / Plaid", "Striped", "Printed", "Self Design", "To be confirmed"],
            "fit_type": ["Standard Regular Fit", "Custom Slim Fit", "Straight Fit", "To be confirmed"],
            "occasion": ["Casual", "Formal", "Party", "Semi-Formal", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Men's [Fabric] [Pattern] [Shirt Type] with [Collar Type] ([Color])",
        "bullet_heading_guidance": ["WOVEN [FABRIC] FABRICATION", "COLLAR & CUFF TAILORING", "COLORFAST & DURABLE", "SMART CASUAL STYLING", "CHEST SIZING & CARE"],
        "supported": True,
        "excluded": False
    },
    "men_tshirt": {
        "profile_id": "men_tshirt",
        "display_name": "Men's T-Shirts & Polos",
        "category_group": "Men Western Wear",
        "amazon_item_type_keyword": "mens-t-shirts",
        "flipkart_vertical": "T-Shirt",
        "meesho_category_hint": "Men Western / T-Shirts",
        "target_gender_options": ["Men"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["fabric", "tshirt_type", "neckline", "sleeve", "pattern", "care_label", "sizes", "color"],
        "optional_verified_fields": ["fit_type", "print_placement", "pocket_count", "occasion"],
        "controlled_attributes": {
            "fabric": ["100% Pure Cotton", "Cotton Blend", "Polyester Blend", "Pique Cotton", "Jersey Knit", "To be confirmed"],
            "tshirt_type": ["Round Neck T-Shirt", "Polo T-Shirt", "Henley T-Shirt", "Oversized T-Shirt", "V-Neck T-Shirt", "To be confirmed"],
            "neckline": ["Round Neck", "Polo Collar", "Henley Neck", "V-Neck", "Hooded", "To be confirmed"],
            "sleeve": ["Half Sleeve", "Full Sleeve", "Sleeveless", "Short Sleeve", "To be confirmed"],
            "pattern": ["Solid", "Typography / Graphic Print", "Striped", "Colorblock", "All-Over Print", "To be confirmed"],
            "fit_type": ["Regular Fit", "Oversized Fit", "Slim Fit", "Relaxed Fit", "To be confirmed"],
            "occasion": ["Casual", "Sports / Active", "Lounge", "Party", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Men's [Fabric] [Pattern] [T-Shirt Type] ([Color])",
        "bullet_heading_guidance": ["KNITTED [FABRIC]", "COLLAR & SLEEVE FINISH", "COLORFAST & DURABLE", "DAILY CASUAL ATTIRE", "SIZE SPECIFICATIONS & CARE"],
        "supported": True,
        "excluded": False
    },
    "men_bottomwear": {
        "profile_id": "men_bottomwear",
        "display_name": "Men's Bottomwear (Jeans / Trousers / Chinos / Joggers)",
        "category_group": "Men Western Wear",
        "amazon_item_type_keyword": "mens-trousers-and-jeans",
        "flipkart_vertical": "Trouser / Jeans",
        "meesho_category_hint": "Men Western / Bottomwear",
        "target_gender_options": ["Men"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["bottom_type", "fabric", "pattern", "waist_type", "length_or_inseam", "closure_type", "care_label", "sizes", "color"],
        "optional_verified_fields": ["rise", "fit_type", "pocket_count", "stretch", "belt_loops", "occasion"],
        "controlled_attributes": {
            "bottom_type": ["Jeans", "Chinos", "Formal Trousers", "Casual Trousers", "Joggers", "Cargo Pants", "To be confirmed"],
            "fabric": ["Cotton Denim", "Cotton Twill", "Poly Viscose", "Pure Cotton", "Linen Blend", "To be confirmed"],
            "pattern": ["Solid", "Checked", "Textured", "Printed", "To be confirmed"],
            "waist_type": ["Fixed Waistband with Belt Loops", "Elasticated Drawstring Waistband", "Semi-Elasticated", "To be confirmed"],
            "rise": ["Mid Rise", "High Rise", "Low Rise", "To be confirmed"],
            "length": ["Full Length", "Ankle Length", "Cropped", "To be confirmed"],
            "fit_type": ["Regular Straight Fit", "Slim Fit", "Tapered Fit", "Relaxed Fit", "To be confirmed"],
            "occasion": ["Casual", "Formal", "Office Wear", "Semi-Formal", "Party", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Men's [Fabric] [Pattern] [Bottom Type] with [Closure/Waist Detail] ([Color])",
        "bullet_heading_guidance": ["STURDY [FABRIC] FABRIC", "WAIST & POCKET CONSTRUCTION", "COLORFAST & DURABLE", "WARDROBE VERSATILITY", "WAIST & INSEAM SIZING"],
        "supported": True,
        "excluded": False
    },
    "women_bottomwear": {
        "profile_id": "women_bottomwear",
        "display_name": "Women's Bottomwear (Jeans / Trousers / Palazzo / Skirts)",
        "category_group": "Women Western & Ethnic Bottoms",
        "amazon_item_type_keyword": "womens-bottomwear",
        "flipkart_vertical": "Women Bottomwear",
        "meesho_category_hint": "Women Western / Bottomwear",
        "target_gender_options": ["Women", "Girls"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["bottom_type", "fabric", "pattern", "waist_type", "length_or_inseam", "closure_type", "care_label", "sizes", "color"],
        "optional_verified_fields": ["rise", "fit_type", "pocket_count", "stretch", "slits", "occasion"],
        "controlled_attributes": {
            "bottom_type": ["Palazzo", "Trousers / Pants", "Jeans", "Leggings", "Culottes", "Skirts", "Jeggings", "To be confirmed"],
            "fabric": ["Pure Cotton", "Rayon", "Cotton Denim", "Poly Viscose", "Lycra Cotton Blend", "Crepe", "To be confirmed"],
            "pattern": ["Solid", "Printed", "Striped", "Floral Print", "Embroidered", "To be confirmed"],
            "waist_type": ["Elasticated Waistband", "Drawstring Waistband", "Fixed Waist with Zipper", "Semi-Elasticated", "To be confirmed"],
            "rise": ["High Rise", "Mid Rise", "Low Rise", "To be confirmed"],
            "length": ["Ankle Length", "Full Length", "Calf Length / Cropped", "Knee Length", "To be confirmed"],
            "fit_type": ["Flared Wide Leg", "Straight Fit", "Slim Fit", "Relaxed Fit", "Skinny Fit", "To be confirmed"],
            "occasion": ["Casual", "Ethnic & Festive", "Office Wear", "Party", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Women's [Fabric] [Pattern] [Bottom Type] with [Waist/Fit Detail] ([Color])",
        "bullet_heading_guidance": ["FABRIC COMPOSITION", "WAISTBAND & SILHOUETTE", "COLORFAST & DURABLE", "ETHNIC & CASUAL PAIRING", "WAIST SIZING & LENGTH"],
        "supported": True,
        "excluded": False
    },
    "men_ethnic": {
        "profile_id": "men_ethnic",
        "display_name": "Men's Ethnic Wear (Kurta / Pyjama / Nehru Jacket / Set)",
        "category_group": "Men Ethnic Wear",
        "amazon_item_type_keyword": "mens-ethnic-wear",
        "flipkart_vertical": "Men Ethnic Wear",
        "meesho_category_hint": "Men Ethnic / Kurtas",
        "target_gender_options": ["Men"],
        "age_group_options": ["Adults", "Teens"],
        "required_verified_fields": ["garment_type", "fabric", "neckline_or_collar", "sleeve", "length", "pattern", "care_label", "sizes", "color"],
        "optional_verified_fields": ["closure_type", "pocket_count", "package_contents", "bottom_included", "occasion"],
        "controlled_attributes": {
            "garment_type": ["Kurta", "Kurta Pyjama Set", "Nehru Jacket", "Sherwani", "Short Kurta", "Pathani Suit", "To be confirmed"],
            "fabric": ["Pure Cotton", "Cotton Blend", "Silk Blend", "Chanderi Cotton", "Linen Blend", "Jacquard", "To be confirmed"],
            "neckline_or_collar": ["Mandarin Collar", "Collar Neck", "Round Neck with Placket", "Stand Collar", "To be confirmed"],
            "sleeve": ["Full Sleeve", "Roll-Up Sleeve", "Half Sleeve", "Sleeveless", "To be confirmed"],
            "length": ["Knee Length", "Calf Length", "Short / Hip Length", "Ankle Length", "To be confirmed"],
            "pattern": ["Solid", "Embroidered", "Printed", "Self Design", "Jacquard Weave", "To be confirmed"],
            "occasion": ["Festive", "Wedding", "Casual & Festive", "Party", "Puja / Traditional", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] Men's [Fabric] [Pattern] [Garment Type] with [Collar/Sleeve Detail] ([Color])",
        "bullet_heading_guidance": ["ETHNIC [FABRIC] FABRIC", "TRADITIONAL COLLAR & PLACKET", "COLORFAST & DURABLE", "FESTIVE & WEDDING OCCASIONS", "SIZE CHART & MAINTENANCE"],
        "supported": True,
        "excluded": False
    },
    "kidswear": {
        "profile_id": "kidswear",
        "display_name": "Kidswear (Boys & Girls Garments / Sets)",
        "category_group": "Kidswear",
        "amazon_item_type_keyword": "kids-apparel",
        "flipkart_vertical": "Kids Apparel",
        "meesho_category_hint": "Kids / Clothing Sets",
        "target_gender_options": ["Boys", "Girls", "Unisex Kids"],
        "age_group_options": ["0-2 Years (Infant/Toddler)", "2-6 Years", "6-12 Years", "12-16 Years (Teens)"],
        "required_verified_fields": ["target_gender", "age_group", "garment_type", "fabric", "pattern", "care_label", "sizes", "color"],
        "optional_verified_fields": ["sleeve", "neckline_or_collar", "package_contents", "closure_type", "length", "set_count", "character_print", "occasion"],
        "controlled_attributes": {
            "target_gender": ["Boys", "Girls", "Unisex Kids", "To be confirmed"],
            "age_group": ["0-2 Years", "2-6 Years", "6-12 Years", "12-16 Years", "To be confirmed"],
            "garment_type": ["Clothing Set (Top & Bottom)", "Dress / Frock", "Kurta Set", "T-Shirt & Shorts", "Shirt", "Trousers / Jeans", "To be confirmed"],
            "fabric": ["100% Pure Cotton", "Cotton Blend", "Hosiery Cotton", "Denim", "Polyester Blend", "To be confirmed"],
            "sleeve": ["Short Sleeve", "Full Sleeve", "Sleeveless", "3/4 Sleeve", "To be confirmed"],
            "neckline_or_collar": ["Round Neck", "Polo Collar", "Collar Neck", "Mandarin Collar", "To be confirmed"],
            "pattern": ["Printed", "Solid", "Graphic Print", "Striped", "Floral Print", "Embroidered", "To be confirmed"],
            "occasion": ["Casual", "Party / Birthday", "Festive", "Playwear", "To be confirmed"],
            "size_class": ["Age-Based (e.g. 3-4Y, 5-6Y)", "Standard Numeric (24, 26, 28)", "Standard S/M/L", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] [Target Gender] [Age Group] [Fabric] [Pattern] [Garment Type] ([Color])",
        "bullet_heading_guidance": ["NATURAL [FABRIC] KNIT", "CHILD-SAFE CLOSURE & SEAMS", "COLORFAST & DURABLE", "PLAYWEAR & OCCASIONS", "AGE-BASED SIZING & CARE"],
        "supported": True,
        "excluded": False
    },
    "footwear": {
        "profile_id": "footwear",
        "display_name": "Footwear (Casual / Formal / Sports / Ethnic)",
        "category_group": "Footwear",
        "amazon_item_type_keyword": "casual-shoes",
        "flipkart_vertical": "Footwear",
        "meesho_category_hint": "Men/Women Footwear / Casual Shoes",
        "target_gender_options": ["Men", "Women", "Unisex Adults", "Kids"],
        "age_group_options": ["Adults", "Kids"],
        "required_verified_fields": ["footwear_type", "material", "closure_type", "sole_material", "toe_shape", "heel_type_or_flat", "care_label", "sizes", "color"],
        "optional_verified_fields": ["heel_height", "occasion", "pattern", "pair_count", "water_resistance", "size_system"],
        "controlled_attributes": {
            "footwear_type": ["Sneakers", "Loafers", "Formal Shoes", "Flat Sandals", "Juttis / Mojaris", "Slip-On Shoes", "Mules", "To be confirmed"],
            "material": ["Synthetic Leather / PU", "Mesh Fabric", "Canvas", "Genuine Leather", "Textile Fabric", "EVA", "To be confirmed"],
            "closure_type": ["Lace-Up", "Slip-On", "Velcro Straps", "Buckle Strap", "Elasticated", "To be confirmed"],
            "sole_material": ["TPR (Thermoplastic Rubber)", "EVA Sole", "Rubber Sole", "PVC Sole", "PU Sole", "To be confirmed"],
            "toe_shape": ["Round Toe", "Pointed Toe", "Square Toe", "Open Toe", "To be confirmed"],
            "heel_type": ["Flat Sole", "Block Heel", "Wedge Heel", "Platform Sole", "To be confirmed"],
            "occasion": ["Casual", "Formal / Office", "Ethnic & Festive", "Party", "Sports / Training", "To be confirmed"],
            "size_system": ["Indian / UK Sizing (e.g. UK 6 to UK 11)", "Euro Sizing (39-44)", "US Sizing", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] [Target Gender] [Material] [Footwear Type] with [Closure Type] and [Sole Material] ([Color])",
        "bullet_heading_guidance": ["UPPER MATERIAL SPECIFICATION", "OUTSOLE & TOE CONSTRUCTION", "MAINTENANCE & DURABILITY", "CASUAL & FORMAL WEAR", "UK/IND SIZING GUIDE"],
        "supported": True,
        "excluded": False
    },
    "home_textiles": {
        "profile_id": "home_textiles",
        "display_name": "Home Textiles & Furnishing (Bedsheets / Curtains / Cushions / Towels)",
        "category_group": "Home & Furnishing",
        "amazon_item_type_keyword": "home-furnishing",
        "flipkart_vertical": "Home Furnishing",
        "meesho_category_hint": "Home & Kitchen / Bedding",
        "target_gender_options": ["Unisex Home"],
        "age_group_options": ["All"],
        "required_verified_fields": ["product_type", "material", "pattern", "dimensions", "package_contents", "care_label", "color"],
        "optional_verified_fields": ["thread_count", "filling_material", "compatible_bed_size", "closure_type", "weave", "occasion"],
        "controlled_attributes": {
            "home_textile_type": ["Bedsheet Set", "Window Curtains", "Door Curtains", "Cushion Covers", "Bath Towels", "Table Runners", "To be confirmed"],
            "material": ["100% Pure Cotton", "Microfiber Glace Cotton", "Polyester Blend", "Linen Blend", "Cotton Terry", "To be confirmed"],
            "pattern": ["Floral Print", "Geometric Print", "Solid", "Abstract Print", "Striped", "Jaipuri Traditional Print", "To be confirmed"],
            "dimensions": ["King Size (108 x 108 in / 274 x 274 cm)", "Queen Size (90 x 100 in / 228 x 254 cm)", "Single Bed (60 x 90 in / 152 x 228 cm)", "Curtain 7 Ft (84 x 48 in)", "Cushion Cover (16 x 16 in)", "To be confirmed"],
            "package_contents": ["1 Bedsheet with 2 Pillow Covers", "1 Bedsheet with 1 Pillow Cover", "Set of 2 Curtains", "Set of 5 Cushion Covers", "Pack of 2 Bath Towels", "To be confirmed"],
            "care": ["Machine Wash Normal", "Gentle Hand Wash", "Cold Wash with Like Colors", "Do Not Bleach", "To be confirmed"]
        },
        "primary_title_template_guidance": "[Brand] [Material] [Pattern] [Product Type] with [Package Contents] ([Dimensions], [Color])",
        "bullet_heading_guidance": ["[MATERIAL] WEAVE SPEC", "ACCURATE DIMENSIONS & FIT", "COLORFAST & DURABLE", "HOME INTERIOR DECOR", "PACKAGE CONTENTS & WASH CARE"],
        "supported": True,
        "excluded": False
    }
}

CATEGORY_MAP = {
    p["category_group"]: p["amazon_item_type_keyword"]
    for p in CATEGORY_PROFILES.values()
}
CATEGORY_MAP["Women Ethnic Wear"] = "kurtas-and-ethnic-tops"
CATEGORY_MAP["Men Western Wear"] = "mens-casual-shirts"
CATEGORY_MAP["Sarees"] = "sarees"
CATEGORY_MAP["Footwear"] = "casual-shoes"
CATEGORY_MAP["Home & Kitchen"] = "home-furnishing"


# Declared Image Roles (Assigned by filename suffix; does not visually verify image content)
IMAGE_ROLES = {
    "_MAIN": "Declared Hero Image",
    "_PT01": "Declared Size Chart Slot",
    "_PT02": "Declared Fabric Specification Slot",
    "_PT03": "Declared Care Guide Slot",
    "_PT04": "Declared Back View Slot",
    "_PT05": "Declared Other Image 5 Slot",
}

CORE_SUFFIXES = ["_MAIN", "_PT01", "_PT02", "_PT03"]


# ──────────────────────────────────────────────
# Pydantic Models (Category-Profile-Driven)
# ──────────────────────────────────────────────

class SellerConfig(BaseModel):
    amazon_quantity: int = 50
    gst_percent: int = 5
    hsn_code: str = "62114200"


class BatchConfig(BaseModel):
    brand: Optional[str] = None
    category: Optional[str] = None
    category_profile: Optional[str] = None
    seller_config: Optional[SellerConfig] = None


class AmazonData(BaseModel):
    title: str
    bullet_points: list[str] = Field(..., min_length=5, max_length=5)
    backend_search_terms: str
    description: str

    @field_validator("title")
    @classmethod
    def validate_title(cls, v: str) -> str:
        if v != TO_BE_CONFIRMED and len(v) > 180:
            raise ValueError(f"Amazon title must be ≤ 180 characters (got {len(v)} chars)")
        if not v.strip():
            raise ValueError("Amazon title cannot be empty")
        return v

    @field_validator("backend_search_terms")
    @classmethod
    def validate_bst(cls, v: str) -> str:
        if v != TO_BE_CONFIRMED:
            b = len(v.encode("utf-8"))
            if b > 240:
                raise ValueError(f"Amazon backend search terms must be ≤ 240 bytes (got {b} bytes)")
        if not v.strip():
            raise ValueError("Amazon backend search terms cannot be empty")
        return v


class FlipkartData(BaseModel):
    title: str
    search_keywords: str
    description: str
    attributes: dict[str, Any] = Field(default_factory=dict)
    # Legacy optional flat fields for backward compatibility
    fabric: Optional[str] = None
    kurta_type: Optional[str] = None
    neck: Optional[str] = None
    sleeve: Optional[str] = None
    length_type: Optional[str] = None
    pattern: Optional[str] = None
    occasion: Optional[str] = None


class MeeshoData(BaseModel):
    title: str
    hinglish_hook_description: str
    english_hook_description: str
    highlights: list[str] = Field(..., min_length=4, max_length=4)

    @field_validator("title")
    @classmethod
    def validate_title(cls, v: str) -> str:
        if v != TO_BE_CONFIRMED and len(v) > 60:
            raise ValueError(f"Meesho title must be ≤ 60 characters (got {len(v)} chars)")
        if not v.strip():
            raise ValueError("Meesho title cannot be empty")
        return v


class AlternateAmazonData(BaseModel):
    title: str
    bullet_points: list[str] = Field(..., min_length=5, max_length=5)
    backend_search_terms: str
    description: str

    @field_validator("title")
    @classmethod
    def validate_title(cls, v: str) -> str:
        if v != TO_BE_CONFIRMED and len(v) > 180:
            raise ValueError(f"Alternate Amazon title must be ≤ 180 characters (got {len(v)} chars)")
        return v

    @field_validator("backend_search_terms")
    @classmethod
    def validate_bst(cls, v: str) -> str:
        if v != TO_BE_CONFIRMED:
            b = len(v.encode("utf-8"))
            if b > 240:
                raise ValueError(f"Alternate Amazon backend search terms must be ≤ 240 bytes (got {b} bytes)")
        return v


class AlternateFlipkartData(BaseModel):
    title: str
    search_keywords: str
    description: str
    attributes: dict[str, Any] = Field(default_factory=dict)


class AlternateMeeshoData(BaseModel):
    title: str
    hinglish_hook_description: str
    english_hook_description: str
    highlights: list[str] = Field(..., min_length=4, max_length=4)

    @field_validator("title")
    @classmethod
    def validate_title(cls, v: str) -> str:
        if v != TO_BE_CONFIRMED and len(v) > 60:
            raise ValueError(f"Alternate Meesho title must be ≤ 60 characters (got {len(v)} chars)")
        return v


class AlternateVariant(BaseModel):
    variant_id: str
    angle_theme: str
    amazon: AlternateAmazonData
    flipkart: AlternateFlipkartData
    meesho: AlternateMeeshoData


class SKUItem(BaseModel):
    sku_id: str
    category_profile: str = "women_ethnic_kurta"
    brand: str
    product_type: str
    color: str
    category: str = "Women Ethnic Wear"
    sizes: str
    mrp: float
    meesho_price: float
    seller_config: SellerConfig
    verified: dict[str, Any] = Field(default_factory=dict)
    amazon: AmazonData
    flipkart: FlipkartData
    meesho: MeeshoData
    alternates: list[AlternateVariant] = Field(default_factory=list, max_length=5)

    @field_validator("alternates")
    @classmethod
    def validate_alternates(cls, v: list[AlternateVariant]) -> list[AlternateVariant]:
        if v and len(v) != 5:
            raise ValueError(f"When alternates are provided, exactly 5 variants (V1-V5) are required (got {len(v)})")
        return v


# ──────────────────────────────────────────────
# Title Quality, Deduplication, & Claims Validators
# ──────────────────────────────────────────────

COMMON_PRODUCT_TYPES = [
    "kurta", "kurti", "tunic", "top", "dress", "anarkali", "suit", "gown", "set", "shirt", "t-shirt", "tshirt",
    "saree", "jeans", "trouser", "trousers", "chinos", "joggers", "palazzo", "leggings", "skirt", "shoes",
    "sneakers", "sandals", "bedsheet", "curtains", "towel", "cushion"
]

def validate_amazon_title_quality(title: str, brand: str) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not title.strip() or title == TO_BE_CONFIRMED:
        return errors, warnings

    if brand and brand != TO_BE_CONFIRMED:
        brand_clean = re.escape(brand.strip())
        brand_matches = list(re.finditer(r"\b" + brand_clean + r"\b", title, re.IGNORECASE))
        if len(brand_matches) > 1:
            errors.append(f"Amazon title contains the brand name '{brand}' {len(brand_matches)} times. Brand must only appear once at the beginning.")

    tokens = re.findall(r"[A-Za-z0-9\-]+", title)
    for i in range(len(tokens) - 1):
        w1 = tokens[i].lower()
        w2 = tokens[i + 1].lower()
        if w1 == w2 and len(w1) > 1:
            errors.append(f"Amazon title contains malformed adjacent word repetition: '{tokens[i]} {tokens[i+1]}'.")
            break

    redundant_phrase_patterns = [
        (r"\bfloral\s+print\s+printed\b", "Floral Print Printed"),
        (r"\bpure\s+cotton\s+cotton\b", "Pure Cotton Cotton"),
        (r"\bindigo\s+blue\s+blue\b", "Indigo Blue Blue"),
        (r"\ba-line\s+a\s+line\b", "A-Line A line"),
        (r"\ba\s+line\s+a-line\b", "A line A-Line"),
    ]
    for pat, label in redundant_phrase_patterns:
        if re.search(pat, title, re.IGNORECASE):
            errors.append(f"Amazon title contains redundant phrase stack: '{label}'.")

    title_lower = title.lower()
    found_types = []
    for pt in COMMON_PRODUCT_TYPES:
        if re.search(r"\b" + re.escape(pt) + r"(s)?\b", title_lower):
            found_types.append(pt)
    
    if ("kurta" in found_types and "kurti" in found_types):
        warnings.append("Amazon title contains both 'kurta' and 'kurti'. Use a single consistent product-type term.")
    if len(set(found_types)) >= 3:
        warnings.append(f"Amazon title contains multiple product-type descriptors ({', '.join(found_types)}). Avoid excessive keyword stacking.")

    pipe_count = title.count("|")
    dash_count = title.count(" - ")
    if pipe_count > 2 or (pipe_count >= 1 and dash_count >= 2):
        warnings.append("Amazon title contains multiple separators/delimiters ('|' or ' - '). Keep title structure natural and customer-readable.")

    return errors, warnings


UNVERIFIED_HARD_CLAIM_PATTERNS = [
    (r"\b(ultra-?)?breathable\b", "breathable", "fabric breathability"),
    (r"\bairy\s+(weave|fabric|cotton)\b", "airy fabric", "fabric ventilation"),
    (r"\bcooling(\s+comfort)?\b", "cooling", "cooling performance"),
    (r"\bsweat-?absorb(ent|ing)?\b", "sweat-absorbent", "sweat absorption"),
    (r"\bquick-?dry\b", "quick-dry", "quick-drying claim"),
    (r"\bnon-?sticky\b", "non-sticky", "tactile performance"),
    (r"\blightweight\b", "lightweight", "fabric weight claim"),
    (r"\bskin-?friendly\b", "skin-friendly", "hypoallergenic/dermatological claim"),
    (r"\bwrinkle-?free\b", "wrinkle-free", "crease resistance claim"),
    (r"\bwrinkle-?resistant\b", "wrinkle-resistant", "crease resistance claim"),
    (r"\bzero\s+fad(e|ing)\b", "zero fading", "unsupported color guarantee"),
    (r"\bzero\s+bleeding\b", "zero bleeding", "unsupported color guarantee"),
    (r"\bno\s+(fading|bleeding)\b", "no fading/bleeding", "unsupported color guarantee"),
    (r"\bpre-?shrunk\b", "pre-shrunk", "unsupported textile processing claim"),
    (r"\banti-?pilling\b", "anti-pilling", "unsupported textile durability claim"),
    (r"\bsuperior(\s+quality)?\b", "superior quality", "unverifiable superiority claim"),
    (r"\bbest\s+quality\b", "best quality", "unverifiable absolute claim"),
    (r"\b(100%\s+)?guaranteed\b", "guaranteed", "unsupported guarantee claim"),
    (r"\bluxury\b", "luxury", "unverifiable luxury claim"),
]

UNVERIFIED_WARNING_CLAIM_PATTERNS = [
    (r"\bcomfortable\s+fit\b", "comfortable fit", "use verified sizing measurements instead"),
    (r"\bcomfort\s+fit\b", "comfort fit", "use verified silhouette or cut instead"),
    (r"\bperfect\s+fit\b", "perfect fit", "fit is subjective; specify size dimensions"),
    (r"\brelaxed\s+fit\b", "relaxed fit", "use verified standard fit or silhouette"),
    (r"\beasy-?to-?wear(\s+fit)?\b", "easy-to-wear", "use styling suggestion instead"),
    (r"\bflattering\s+fit\b", "flattering fit", "subjective fit claim; state silhouette"),
    (r"\btailored\s+fit\b", "tailored fit", "use verified silhouette (e.g. Straight/A-line)"),
    (r"\bsmart\s+fit\b", "smart fit", "use standard size specifications"),
    (r"\bideal\s+fit\b", "ideal fit", "use verified size dimensions"),
    (r"\b(ultra-?)?soft\s+(feel|cotton|fabric|touch)\b", "soft feel", "tactile feel is subjective"),
    (r"\bpremium\s+(fabric|cotton|quality)\b", "premium fabric", "describe thread count/weave factually"),
    (r"\brich\s+(fabric|look)\b", "rich look", "describe visual color/print factually"),
]

def validate_truthfulness_and_claims(
    field_path: str,
    text: str,
    category_profile: str = "women_ethnic_kurta",
    verified_data: Optional[dict] = None
) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not text or not text.strip() or text == TO_BE_CONFIRMED:
        return errors, warnings

    text_lower = text.lower()
    verified_data = verified_data or {}

    # 1. Universal hard performance/guarantee claims
    for pattern, term, desc in UNVERIFIED_HARD_CLAIM_PATTERNS:
        if re.search(pattern, text_lower):
            errors.append(
                f"Field '{field_path}' contains unverified performance/guarantee claim ('{term}'). "
                f"Remove unverified {desc} unless explicitly documented in verified product records."
            )

    # 2. Universal subjective fit/comfort warnings
    for pattern, term, suggestion in UNVERIFIED_WARNING_CLAIM_PATTERNS:
        if re.search(pattern, text_lower):
            warnings.append(
                f"Field '{field_path}' contains unverified claim/descriptor ('{term}'). Suggestion: {suggestion}."
            )

    # 3. Category-Specific Claim Boundaries
    if category_profile == "footwear":
        fw_forbidden = [
            (r"\banti-?slip\b", "anti-slip", "sole traction verification"),
            (r"\bcushion(ed|ing)?\b", "cushioning", "insole cushioning certification"),
            (r"\barch\s+support\b", "arch support", "orthopedic support verification"),
            (r"\bmemory\s+foam\b", "memory foam", "verified insole material"),
            (r"\bwater-?proof\b", "waterproof", "waterproofing laboratory testing"),
        ]
        for pat, term, req in fw_forbidden:
            if re.search(pat, text_lower) and term not in str(verified_data).lower():
                errors.append(f"Field '{field_path}' contains unverified footwear claim ('{term}'). Requires {req}.")

    elif category_profile == "saree":
        saree_forbidden = [
            (r"\bhandloom\b", "handloom", "government handloom certification"),
            (r"\bbanarasi\b", "Banarasi", "GI tag / certified Banarasi origin"),
            (r"\bkanjivaram\b", "Kanjivaram", "certified Kanchipuram silk origin"),
            (r"\bpure\s+silk\b", "pure silk", "Silk Mark certification"),
            (r"\bartisanal\b", "artisanal", "verified artisan provenance"),
        ]
        for pat, term, req in saree_forbidden:
            if re.search(pat, text_lower) and term.lower() not in str(verified_data).lower():
                errors.append(f"Field '{field_path}' contains unverified saree claim ('{term}'). Requires {req}.")

    elif category_profile == "home_textiles":
        ht_forbidden = [
            (r"\b\d{3,4}\s*(tc|thread\s*count)\b", "thread count", "laboratory thread count test"),
            (r"\b\d{3,4}\s*gsm\b", "GSM", "verified fabric grammage"),
            (r"\bhypoallergenic\b", "hypoallergenic", "dermatological allergy testing"),
            (r"\bstain-?resistant\b", "stain-resistant", "chemical coating verification"),
        ]
        for pat, term, req in ht_forbidden:
            if re.search(pat, text_lower) and term.lower() not in str(verified_data).lower():
                errors.append(f"Field '{field_path}' contains unverified home textile claim ('{term}'). Requires {req}.")

    elif category_profile == "kidswear":
        kids_forbidden = [
            (r"\bskin-?safe\b", "skin-safe", "pediatric dermatological test"),
            (r"\bgentle\s+on\s+skin\b", "gentle on skin", "fabric safety certification"),
            (r"\b100%\s+organic\b", "organic", "GOTS organic certification"),
        ]
        for pat, term, req in kids_forbidden:
            if re.search(pat, text_lower) and term.lower() not in str(verified_data).lower():
                errors.append(f"Field '{field_path}' contains unverified kidswear claim ('{term}'). Requires {req}.")

    elif category_profile in ["men_shirt", "men_bottomwear", "women_bottomwear"]:
        if "fit_type" not in verified_data:
            fit_pats = [
                (r"\bslim\s+fit\b", "slim fit"),
                (r"\btailored\s+fit\b", "tailored fit"),
                (r"\bcomfort\s+fit\b", "comfort fit"),
            ]
            for pat, term in fit_pats:
                if re.search(pat, text_lower):
                    warnings.append(f"Field '{field_path}' uses fit term '{term}' without verified fit_type in product record.")

    return errors, warnings


def validate_amazon_backend_search_terms(title: str, bst: str, brand: str) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not bst.strip() or bst == TO_BE_CONFIRMED:
        return errors, warnings

    try:
        bst.encode("ascii")
    except UnicodeEncodeError:
        errors.append("Amazon backend search terms must contain ASCII characters only.")

    if re.search(r"[A-Z]", bst):
        errors.append("Amazon backend search terms must be lowercase only (uppercase characters found).")

    invalid_chars = set(re.findall(r"[^a-z0-9\s]", bst))
    if invalid_chars:
        errors.append(f"Amazon backend search terms must not contain punctuation or commas (found: {', '.join(sorted(invalid_chars))}). Use single spaces only.")

    if re.search(r"\s{2,}", bst):
        errors.append("Amazon backend search terms must not contain repeated whitespace. Use single spaces between terms.")

    bst_lower = bst.lower()
    bst_tokens = bst_lower.split()
    brand_tokens = [t.lower() for t in re.findall(r"[a-z0-9]+", brand) if len(t) >= 2]
    for bt in brand_tokens:
        if bt in bst_tokens or bt in bst_lower:
            errors.append(f"Amazon backend search terms must not contain the brand name ('{brand}' / '{bt}').")
            break

    stopwords = {"and", "or", "with", "for", "in", "of", "a", "an", "the", "to", "by", "on", "at", "from", "is", "it", "as"}
    title_tokens = set(re.findall(r"[a-z0-9]+", title.lower())) - stopwords

    exact_dupes = []
    for word in bst_tokens:
        clean_w = re.sub(r"[^a-z0-9]", "", word)
        if not clean_w or clean_w in stopwords:
            continue
        if clean_w in title_tokens:
            exact_dupes.append(clean_w)

    if exact_dupes:
        warnings.append(f"Amazon backend search terms contain word(s) already in the title: {', '.join(sorted(set(exact_dupes)))}. Consider replacing with non-title search keywords.")

    return errors, warnings


def validate_bullet_3_safety(bullet_3: str) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not bullet_3.strip() or bullet_3 == TO_BE_CONFIRMED:
        return errors, warnings

    if not bullet_3.startswith("COLORFAST & DURABLE:"):
        errors.append("Bullet 3 must start with the standard heading 'COLORFAST & DURABLE:'.")

    return errors, warnings


# ──────────────────────────────────────────────
# Category Profile Validator & Normalizer
# ──────────────────────────────────────────────

def normalize_sku_profile_and_attributes(sku_dict: dict[str, Any], batch_config: Optional[BatchConfig] = None) -> tuple[dict[str, Any], list[str]]:
    """
    Normalizes SKU dictionary:
    - Injects batch_config defaults (brand, category, seller_config, category_profile).
    - Infers category_profile if missing for legacy Women Ethnic Wear.
    - Normalizes flat Flipkart fields into flipkart.attributes.
    """
    warnings: list[str] = []
    d = dict(sku_dict)

    if batch_config:
        if "brand" not in d and batch_config.brand:
            d["brand"] = batch_config.brand
        if "category" not in d and batch_config.category:
            d["category"] = batch_config.category
        if "category_profile" not in d and batch_config.category_profile:
            d["category_profile"] = batch_config.category_profile
        if "seller_config" not in d and batch_config.seller_config:
            d["seller_config"] = batch_config.seller_config.model_dump()

    # Profile Inference & Legacy Compatibility
    sk_id = d.get("sku_id", "SKU")
    if not d.get("category_profile"):
        cat = d.get("category", "")
        if cat in ["Women Ethnic Wear", "kurtas-and-ethnic-tops"] or "kurti" in str(d).lower():
            d["category_profile"] = "women_ethnic_kurta"
            warnings.append(f"{sk_id}: category_profile was inferred as 'women_ethnic_kurta' for legacy compatibility. Include it explicitly in future payloads.")
        else:
            d["category_profile"] = "women_ethnic_kurta"
            warnings.append(f"{sk_id}: category_profile was defaulted to 'women_ethnic_kurta'.")

    # Normalize Flipkart attributes
    fk = d.get("flipkart", {})
    if isinstance(fk, dict):
        attrs = dict(fk.get("attributes", {}))
        # Migrate flat legacy fields
        for legacy_k in ["fabric", "kurta_type", "neck", "sleeve", "length_type", "pattern", "occasion", "saree_type", "blouse_piece", "saree_length", "border_type", "top_type", "bottom_type", "dress_type", "shirt_type", "collar_type", "tshirt_type", "waist_type", "rise", "garment_type", "target_gender", "age_group", "footwear_type", "material", "sole_material", "closure_type", "toe_shape", "heel_type", "home_textile_type", "dimensions", "package_contents"]:
            if legacy_k in fk and fk[legacy_k] is not None and legacy_k not in attrs:
                attrs[legacy_k] = fk[legacy_k]
        fk["attributes"] = attrs
        d["flipkart"] = fk

    # Normalize verified facts
    if "verified" not in d or not isinstance(d.get("verified"), dict):
        d["verified"] = {}

    return d, warnings


def validate_sku_category_profile(sku_item: SKUItem) -> tuple[list[str], list[str]]:
    """
    Validates a parsed SKUItem against its selected Category Profile:
    - Rejects excluded profiles (blouses, lingerie/innerwear).
    - Enforces presence of required verified facts.
    - Validates controlled attributes against allowed values.
    """
    errors: list[str] = []
    warnings: list[str] = []

    profile_id = sku_item.category_profile
    sk_id = sku_item.sku_id

    # Excluded Profiles Check
    if profile_id in EXCLUDED_PROFILES:
        errors.append(f"SKU '{sk_id}': Category profile '{profile_id}' ({EXCLUDED_PROFILES[profile_id]}) is intentionally not supported in Listing Factory v2.0.")
        return errors, warnings

    if profile_id not in CATEGORY_PROFILES:
        errors.append(f"SKU '{sk_id}': Unknown category_profile '{profile_id}'. Supported profiles: {', '.join(sorted(CATEGORY_PROFILES.keys()))}")
        return errors, warnings

    profile = CATEGORY_PROFILES[profile_id]
    verified = sku_item.verified or {}
    fk_attrs = sku_item.flipkart.attributes or {}

    # Pool verified facts and attributes for checking required facts
    combined_facts = {**fk_attrs, **verified}
    # Also check top-level fields
    combined_facts["color"] = sku_item.color
    combined_facts["sizes"] = sku_item.sizes
    combined_facts["product_type"] = sku_item.product_type

    # 1. Enforce Required Verified Fields
    for req_f in profile["required_verified_fields"]:
        val = combined_facts.get(req_f)
        if val is None or (isinstance(val, str) and not val.strip()):
            # Check if it's in flipkart attributes or top-level
            errors.append(f"SKU '{sk_id}' ({profile_id}): Missing required verified fact '{req_f}'.")

    # 2. Validate Controlled Attributes against allowed enums
    controlled = profile.get("controlled_attributes", {})
    for attr_k, allowed_vals in controlled.items():
        if attr_k in fk_attrs:
            actual_val = fk_attrs[attr_k]
            if actual_val != TO_BE_CONFIRMED and actual_val not in allowed_vals:
                # Check case-insensitive / partial match
                matched = any(allowed_val.lower() == str(actual_val).lower() for allowed_val in allowed_vals)
                if not matched:
                    errors.append(
                        f"SKU '{sk_id}' ({profile_id}): Controlled attribute '{attr_k}' has invalid value '{actual_val}'. "
                        f"Allowed values: {', '.join(allowed_vals[:6])}..."
                    )

    return errors, warnings


# ──────────────────────────────────────────────
# Master Excel Workbook Builder (Profile-Aware)
# ──────────────────────────────────────────────

def _style_header_row(ws, col_count: int, header_color: str = "1E293B"):
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, col_count + 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = cell_border
    ws.row_dimensions[1].height = 28


def _auto_width(ws, min_width=14, max_width=50):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def _sku_has_unconfirmed_fields(sku: SKUItem) -> bool:
    for val in [sku.product_type, sku.color, sku.sizes, sku.amazon.title, sku.amazon.backend_search_terms, sku.flipkart.title, sku.meesho.title]:
        if str(val).strip().lower() == TO_BE_CONFIRMED.lower():
            return True
    for v in (sku.flipkart.attributes or {}).values():
        if str(v).strip().lower() == TO_BE_CONFIRMED.lower():
            return True
    return False


def _build_master_summary(
    wb: Workbook,
    skus: list[SKUItem],
    image_map: dict[str, list[str]],
    validation_status: str = "✅ Pass"
):
    ws = wb.active
    ws.title = "Master_Summary"
    headers = [
        "SKU ID", "Brand", "Category Profile", "Category", "Product Type", "Color", "Key Attributes / Fabric", "Sizes / Dimensions",
        "Amazon Title Preview", "Flipkart Title Preview", "Meesho Hinglish Hook Preview",
        "Core Images Found", "Core Coverage", "Validation Status", "Review Flags", "Package Readiness",
        "Status Scope / Meaning"
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for sku in skus:
        files = image_map.get(sku.sku_id, [])
        core_found = sum(
            1 for suffix in CORE_SUFFIXES
            if any(suffix.lower() in fn.lower() for fn in files)
        )
        coverage_pct = f"{int((core_found / 4) * 100)}% ({core_found}/4 Core)"
        has_tbc = _sku_has_unconfirmed_fields(sku)
        review_flags = "⚠️ Has unconfirmed fields" if has_tbc else "—"

        if validation_status == "✅ Pass" and core_found == 4 and not has_tbc:
            readiness = "✅ Structurally Complete – Seller Review Required"
        elif validation_status == "✅ Pass":
            readiness = "⚠️ Warnings – Seller Review Required"
        elif "Warning" in validation_status:
            readiness = "⚠️ Warnings – Seller Review Required"
        else:
            readiness = "❌ Not Ready – Fix Errors First"

        attrs_summary = ", ".join(f"{k}: {v}" for k, v in list((sku.flipkart.attributes or {}).items())[:3])

        ws.append([
            sku.sku_id,
            sku.brand,
            sku.category_profile,
            sku.category,
            sku.product_type,
            sku.color,
            attrs_summary or sku.product_type,
            sku.sizes,
            (sku.amazon.title[:75] + "…") if len(sku.amazon.title) > 75 else sku.amazon.title,
            (sku.flipkart.title[:75] + "…") if len(sku.flipkart.title) > 75 else sku.flipkart.title,
            (sku.meesho.hinglish_hook_description[:75] + "…") if len(sku.meesho.hinglish_hook_description) > 75 else sku.meesho.hinglish_hook_description,
            f"{core_found}/4 Slots",
            coverage_pct,
            validation_status,
            review_flags,
            readiness,
            STRUCTURAL_READINESS_DISCLAIMER
        ])
    _auto_width(ws)


def _build_amazon_tab(wb: Workbook, skus: list[SKUItem], category: str):
    ws = wb.create_sheet("01_Amazon_Bulk_Import")
    headers = [
        "item_sku", "item_name", "brand_name", "category_profile", "feed_product_type", "item_type_keyword",
        "standard_price", "currency", "quantity", "condition_type",
        "main_image_url", "other_image_url1", "other_image_url2", "other_image_url3", "other_image_url4", "other_image_url5",
        "bullet_point1", "bullet_point2", "bullet_point3", "bullet_point4", "bullet_point5",
        "generic_keyword", "item_description", "size_or_dimensions", "color"
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for sku in skus:
        prof = CATEGORY_PROFILES.get(sku.category_profile, {})
        item_type = prof.get("amazon_item_type_keyword") or CATEGORY_MAP.get(category, "apparel")
        feed_type = prof.get("display_name", "Apparel")
        bp = sku.amazon.bullet_points
        ws.append([
            sku.sku_id,
            sku.amazon.title,
            sku.brand,
            sku.category_profile,
            feed_type,
            item_type,
            sku.mrp,
            "INR",
            sku.seller_config.amazon_quantity,
            "New",
            f"{sku.sku_id}_MAIN.jpg",
            f"{sku.sku_id}_PT01.jpg",
            f"{sku.sku_id}_PT02.jpg",
            f"{sku.sku_id}_PT03.jpg",
            f"{sku.sku_id}_PT04.jpg",
            f"{sku.sku_id}_PT05.jpg",
            bp[0], bp[1], bp[2], bp[3], bp[4],
            sku.amazon.backend_search_terms,
            sku.amazon.description,
            sku.sizes,
            sku.color,
        ])
    _auto_width(ws)


def _build_flipkart_tab(wb: Workbook, skus: list[SKUItem]):
    ws = wb.create_sheet("02_Flipkart_Bulk_Import")
    
    # Collect distinct controlled attribute keys from the active SKUs
    attr_keys = []
    for sku in skus:
        prof = CATEGORY_PROFILES.get(sku.category_profile, {})
        controlled = prof.get("controlled_attributes", {})
        for k in controlled.keys():
            if k not in attr_keys:
                attr_keys.append(k)
        for k in (sku.flipkart.attributes or {}).keys():
            if k not in attr_keys:
                attr_keys.append(k)

    if not attr_keys:
        attr_keys = ["fabric", "pattern", "occasion"]

    headers = [
        "Seller SKU ID", "Product Title", "Brand", "Style Code", "Category Profile", "Ideal For", "Size", "Color"
    ] + [k.replace("_", " ").title() for k in attr_keys] + [
        "Net Quantity", "GST (%)", "HSN Code", "Search Keywords",
        "Main Image Name", "Angle 1 Image", "Angle 2 Image", "Angle 3 Image", "Angle 4 Image", "Angle 5 Image",
        "Description"
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for sku in skus:
        prof = CATEGORY_PROFILES.get(sku.category_profile, {})
        ideal_for = prof.get("target_gender_options", ["Unisex"])[0]
        attrs = sku.flipkart.attributes or {}
        attr_values = [str(attrs.get(k, "—")) for k in attr_keys]

        row = [
            sku.sku_id,
            sku.flipkart.title,
            sku.brand,
            sku.sku_id,
            sku.category_profile,
            ideal_for,
            sku.sizes,
            sku.color,
        ] + attr_values + [
            1,
            sku.seller_config.gst_percent,
            sku.seller_config.hsn_code,
            sku.flipkart.search_keywords,
            f"{sku.sku_id}_MAIN.jpg",
            f"{sku.sku_id}_PT01.jpg",
            f"{sku.sku_id}_PT02.jpg",
            f"{sku.sku_id}_PT03.jpg",
            f"{sku.sku_id}_PT04.jpg",
            f"{sku.sku_id}_PT05.jpg",
            sku.flipkart.description,
        ]
        ws.append(row)
    _auto_width(ws)


def _build_meesho_tab(wb: Workbook, skus: list[SKUItem]):
    ws = wb.create_sheet("03_Meesho_Bulk_Import")
    headers = [
        "Product ID / SKU", "Product Name", "Category Profile",
        "Product Description (Hinglish Hook)",
        "English Hook Description",
        "Primary Material / Fabric", "Available Sizes / Dimensions", "Color",
        "Key Highlights", "GST (%)", "HSN Code",
        "Recommended Price (INR)",
        "Primary Image (Hero)",
        "Other Image 1 (Size Chart)",
        "Other Image 2 (Fabric Spec)",
        "Other Image 3 (Care Guide)",
        "Other Image 4 (Back View)",
        "Other Image 5"
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for sku in skus:
        highlights_str = " • ".join(sku.meesho.highlights)
        mat = sku.flipkart.attributes.get("fabric") or sku.flipkart.attributes.get("material") or sku.product_type
        ws.append([
            sku.sku_id,
            sku.meesho.title,
            sku.category_profile,
            sku.meesho.hinglish_hook_description,
            sku.meesho.english_hook_description,
            mat,
            sku.sizes,
            sku.color,
            highlights_str,
            sku.seller_config.gst_percent,
            sku.seller_config.hsn_code,
            sku.meesho_price,
            f"{sku.sku_id}_MAIN.jpg",
            f"{sku.sku_id}_PT01.jpg",
            f"{sku.sku_id}_PT02.jpg",
            f"{sku.sku_id}_PT03.jpg",
            f"{sku.sku_id}_PT04.jpg",
            f"{sku.sku_id}_PT05.jpg",
        ])
    _auto_width(ws)


def build_workbook(
    skus: list[SKUItem],
    category: str,
    image_map: dict[str, list[str]],
    validation_status: str = "✅ Pass"
) -> bytes:
    wb = Workbook()
    _build_master_summary(wb, skus, image_map, validation_status)
    _build_amazon_tab(wb, skus, category)
    _build_flipkart_tab(wb, skus)
    _build_meesho_tab(wb, skus)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_alternates_workbook(skus: list[SKUItem]) -> bytes:
    wb = Workbook()

    # -- Tab 1: Amazon Alternates --
    ws_amz = wb.active
    ws_amz.title = "Amazon_Alternate_Copies"
    headers_amz = [
        "SKU ID", "Brand", "Category Profile", "Variant", "Marketing Angle / Theme",
        "Alternative Title (Amazon)",
        "Bullet 1", "Bullet 2", "Bullet 3", "Bullet 4", "Bullet 5",
        "Alternative Backend Search Terms (<240 bytes)",
        "Alternative Description"
    ]
    ws_amz.append(headers_amz)
    _style_header_row(ws_amz, len(headers_amz), header_color="38BDF8")

    # -- Tab 2: Flipkart Alternates --
    ws_fk = wb.create_sheet("Flipkart_Alternate_Copies")
    headers_fk = [
        "SKU ID", "Brand", "Category Profile", "Variant", "Marketing Angle / Theme",
        "Alternative Flipkart Title",
        "Alternative Search Keywords",
        "Alternative Product Description"
    ]
    ws_fk.append(headers_fk)
    _style_header_row(ws_fk, len(headers_fk), header_color="38BDF8")

    # -- Tab 3: Meesho Alternates --
    ws_me = wb.create_sheet("Meesho_Alternate_Copies")
    headers_me = [
        "SKU ID", "Brand", "Category Profile", "Variant", "Marketing Angle / Theme",
        "Alternative Product Title",
        "Hinglish Hook Description",
        "English Hook Description",
        "Key Highlights"
    ]
    ws_me.append(headers_me)
    _style_header_row(ws_me, len(headers_me), header_color="38BDF8")

    for sku in skus:
        for idx, alt in enumerate(sku.alternates):
            v_tag = alt.variant_id or f"V{idx+1}"
            theme = alt.angle_theme or f"Angle {idx+1}"
            bp = alt.amazon.bullet_points
            ws_amz.append([
                sku.sku_id, sku.brand, sku.category_profile, v_tag, theme,
                alt.amazon.title,
                bp[0], bp[1], bp[2], bp[3], bp[4],
                alt.amazon.backend_search_terms,
                alt.amazon.description,
            ])
            ws_fk.append([
                sku.sku_id, sku.brand, sku.category_profile, v_tag, theme,
                alt.flipkart.title,
                alt.flipkart.search_keywords,
                alt.flipkart.description,
            ])
            ws_me.append([
                sku.sku_id, sku.brand, sku.category_profile, v_tag, theme,
                alt.meesho.title,
                alt.meesho.hinglish_hook_description,
                alt.meesho.english_hook_description,
                " • ".join(alt.meesho.highlights),
            ])

    _auto_width(ws_amz)
    _auto_width(ws_fk)
    _auto_width(ws_me)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# Package Metadata & README Generator
# ──────────────────────────────────────────────

def build_package_metadata(
    client_name: str,
    batch_id: str,
    category: str,
    skus: list[SKUItem],
    schema_version: str = EXPECTED_SCHEMA_VERSION,
) -> dict[str, Any]:
    generated_at = datetime.utcnow().isoformat() + "Z"
    skus_meta = []
    for sku in skus:
        raw_in = json.dumps(sku.model_dump(), sort_keys=True).encode("utf-8")
        in_hash = hashlib.sha256(raw_in).hexdigest()
        out_str = f"{sku.sku_id}|{sku.amazon.title}|{sku.flipkart.title}|{sku.meesho.title}"
        out_hash = hashlib.sha256(out_str.encode("utf-8")).hexdigest()
        skus_meta.append({
            "sku_id": sku.sku_id,
            "category_profile": sku.category_profile,
            "brand": sku.brand,
            "product_type": sku.product_type,
            "color": sku.color,
            "input_hash": f"sha256:{in_hash}",
            "output_hash": f"sha256:{out_hash}"
        })
    return {
        "tool_version": TOOL_VERSION,
        "json_prompt_version": JSON_PROMPT_VERSION,
        "schema_version": schema_version,
        "client_name": client_name,
        "batch_id": batch_id,
        "category": category,
        "generated_at": generated_at,
        "total_skus": len(skus),
        "structural_readiness_disclaimer": STRUCTURAL_READINESS_DISCLAIMER,
        "image_role_disclaimer": IMAGE_ROLE_DISCLAIMER,
        "skus": skus_meta
    }


def generate_readme(client_name: str, batch_id: str, category: str, profile_id: str = "women_ethnic_kurta") -> str:
    prof = CATEGORY_PROFILES.get(profile_id, CATEGORY_PROFILES["women_ethnic_kurta"])
    return f"""================================================================================
  LISTING FACTORY v2.1 -- CLIENT DELIVERY PACKAGE INSTRUCTIONS
================================================================================
Client Name      : {client_name}
Batch ID         : {batch_id}
Category Profile : {prof['display_name']} ({profile_id})
Generated At     : {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC
Tool Version     : {TOOL_VERSION}
JSON Prompt Ver  : {JSON_PROMPT_VERSION}

================================================================================
  STRUCTURAL READINESS & COMPLIANCE NOTICE
================================================================================

  {STRUCTURAL_READINESS_DISCLAIMER}

  Important Operational Caveats:
  • Structural readiness confirms data format and filename rules only.
  • It does NOT guarantee marketplace acceptance, indexing, or live buyability.
  • It does not provide legal, tax, or official classification advice (GST/HSN).
  • All data must be reviewed and confirmed by the seller before portal upload.

================================================================================
  TEMPLATE REFERENCE NOTES
================================================================================

  This package mapping is tailored for:
    • Amazon.in  : '{prof['category_group']}' (Keyword: {prof['amazon_item_type_keyword']})
    • Flipkart   : {prof['flipkart_vertical']} Vertical
    • Meesho     : {prof['meesho_category_hint']}

================================================================================
  IMAGE NAMING & ASSET STRUCTURE
================================================================================

  {IMAGE_ROLE_DISCLAIMER}

  Canonical Image Naming Scheme (Declared Filename Slots):
    • Primary Image (Hero)        : SKU_XX_MAIN.jpg (Declared Hero cutout)
    • Other Image 1 (Size Chart)  : SKU_XX_PT01.jpg (Declared Size Guide)
    • Other Image 2 (Fabric Spec) : SKU_XX_PT02.jpg (Declared Feature / Fabric Spec)
    • Other Image 3 (Care Guide)  : SKU_XX_PT03.jpg (Declared Care / Detail Slot)
    • Other Image 4 (Back View)   : SKU_XX_PT04.jpg (Declared Back / Angle Slot)
    • Other Image 5               : SKU_XX_PT05.jpg (Declared Other Image 5)

  Folder Hierarchy in this ZIP:
    Organized_SKU_Images/
      ├── [SKU_ID]/              ← One subfolder per SKU
      │   ├── [SKU]_MAIN.jpg     ← Declared Primary hero (manual visual check required)
      │   ├── [SKU]_PT01.jpg     ← Declared Size & measurement guide (manual visual check required)
      │   ├── [SKU]_PT02.jpg     ← Declared Feature/Fabric texture (manual visual check required)
      │   ├── [SKU]_PT03.jpg     ← Declared Care & details (manual visual check required)
      │   ├── [SKU]_PT04.jpg     ← Declared Additional angle (manual visual check required)
      │   └── [SKU]_PT05.jpg     ← Declared Other Image 5 (manual visual check required)
      └── Unassigned_Assets/     ← Assets requiring manual prefix assignment

================================================================================
"""


def route_images(
    sku_ids: list[str],
    image_files: list[tuple[str, bytes]],
) -> tuple[dict[str, list[tuple[str, bytes]]], list[tuple[str, bytes]]]:
    matched: dict[str, list[tuple[str, bytes]]] = {sid: [] for sid in sku_ids}
    unassigned: list[tuple[str, bytes]] = []

    for fname, data in image_files:
        base_name = Path(fname).name
        stem = Path(fname).stem
        found = False
        for sid in sku_ids:
            if stem.startswith(sid):
                matched[sid].append((base_name, data))
                found = True
                break
        if not found:
            unassigned.append((base_name, data))

    return matched, unassigned


def build_zip(
    client_name: str,
    batch_id: str,
    category: str,
    skus: list[SKUItem],
    image_files: list[tuple[str, bytes]],
    validation_status: str = "✅ Pass",
    schema_version: str = EXPECTED_SCHEMA_VERSION,
) -> bytes:
    sku_ids = [s.sku_id for s in skus]
    matched_images, unassigned_images = route_images(sku_ids, image_files)
    image_counts = {sid: [fn for fn, _ in files] for sid, files in matched_images.items()}

    primary_profile = skus[0].category_profile if skus else "women_ethnic_kurta"
    master_wb = build_workbook(skus, category, image_counts, validation_status)
    alt_wb = build_alternates_workbook(skus)
    meta_dict = build_package_metadata(client_name, batch_id, category, skus, schema_version)
    readme_txt = generate_readme(client_name, batch_id, category, primary_profile)

    prefix = f"{client_name}_{batch_id}"
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{prefix}/{prefix}_Master_Marketplace_Upload.xlsx", master_wb)
        zf.writestr(f"{prefix}/{prefix}_Alternate_Listing_Copies.xlsx", alt_wb)
        zf.writestr(f"{prefix}/{prefix}_package_metadata.json", json.dumps(meta_dict, indent=2))
        zf.writestr(f"{prefix}/README.txt", readme_txt)

        for sid, files in matched_images.items():
            for fname, data in files:
                zf.writestr(f"{prefix}/Organized_SKU_Images/{sid}/{fname}", data)
        for fname, data in unassigned_images:
            zf.writestr(f"{prefix}/Organized_SKU_Images/Unassigned_Assets/{fname}", data)

    buf.seek(0)
    return buf.getvalue()


# ──────────────────────────────────────────────
# Sample SKUs Data (All 13 Profiles)
# ──────────────────────────────────────────────

def get_sample_skus_for_profile(profile_id: str) -> list[dict[str, Any]]:
    """Returns a valid, 100% compliant sample SKU payload for any supported profile."""
    prof = CATEGORY_PROFILES.get(profile_id, CATEGORY_PROFILES["women_ethnic_kurta"])
    
    samples: dict[str, dict[str, Any]] = {
        "women_ethnic_kurta": {
            "sku_id": "SKU_01",
            "category_profile": "women_ethnic_kurta",
            "brand": "Anvi Fabrics",
            "category": "Women Ethnic Wear",
            "product_type": "Pure Cotton Anarkali Kurti",
            "color": "Navy Blue",
            "sizes": "S, M, L, XL, XXL, 3XL",
            "mrp": 1299.0,
            "meesho_price": 349.0,
            "seller_config": {"amazon_quantity": 50, "gst_percent": 5, "hsn_code": "62114200"},
            "verified": {"fabric": "Pure Cotton", "sleeve": "3/4 Sleeve", "neckline": "Mandarin Neck", "length": "Calf Length", "pattern": "Floral Print", "care_label": "Machine Wash as per Label", "product_type": "Anarkali Kurti", "color": "Navy Blue", "sizes": "S to 3XL"},
            "amazon": {
                "title": "Anvi Fabrics Women's Pure Cotton Anarkali Kurti with Pockets | Navy Blue Daily Wear",
                "bullet_points": [
                    "100% PURE COTTON: Woven from 60s count pure cotton yarn for everyday wear.",
                    "ANARKALI SILHOUETTE WITH POCKETS: Flared construction with functional side pockets.",
                    "COLORFAST & DURABLE: Follow the provided care label to help maintain the fabric's appearance and color.",
                    "VERSATILE STYLING: Pairs with leggings, palazzos, or denim trousers.",
                    "ACCURATE SIZING: Available from size S (36 in chest) to 3XL (46 in chest)."
                ],
                "backend_search_terms": "tunic dress office ethnic top formal printed attire festive apparel clothing",
                "description": "Pure Cotton Anarkali Kurti by Anvi Fabrics tailored with 60s count cotton yarn and dual side pockets."
            },
            "flipkart": {
                "title": "Anvi Fabrics Women Printed Pure Cotton Anarkali Kurta (Blue)",
                "search_keywords": "anarkali kurti, cotton kurta women, casual kurti",
                "description": "Anvi Fabrics printed pure cotton anarkali kurta with mandarin neck and 3/4 sleeves designed for daily ethnic styling.",
                "attributes": {"fabric": "Pure Cotton", "kurta_type": "Anarkali", "neck": "Mandarin Neck", "sleeve": "3/4 Sleeve", "length_type": "Calf Length", "pattern": "Floral Print", "occasion": "Casual & Festive"}
            },
            "meesho": {
                "title": "Pure Cotton Flared Anarkali Kurti with Pockets",
                "hinglish_hook_description": "100% Pure Cotton fabric kurta daily office aur festive wear ke liye. Side pockets aur flared anarkali ghera ke saath.",
                "english_hook_description": "Pure cotton anarkali kurti crafted with functional side pockets and flared silhouette for everyday ethnic wear.",
                "highlights": ["Fabric: 100% Pure Cotton", "Style: Flared Anarkali with Pockets", "Sizes: S to 3XL", "Care: Hand or Machine Wash as per Label"]
            }
        },
        "saree": {
            "sku_id": "SKU_SAREE_01",
            "category_profile": "saree",
            "brand": "Kalarang Sarees",
            "category": "Sarees",
            "product_type": "Georgette Floral Printed Saree",
            "color": "Emerald Green",
            "sizes": "Free Size (5.5m + 0.8m Blouse Piece)",
            "mrp": 1899.0,
            "meesho_price": 499.0,
            "seller_config": {"amazon_quantity": 40, "gst_percent": 5, "hsn_code": "54075200"},
            "verified": {"fabric": "Georgette", "pattern": "Floral Print", "saree_length": "5.5 m", "blouse_piece_included": "Unstitched Blouse Piece", "care_label": "Dry Clean Recommended or Gentle Hand Wash", "color": "Emerald Green", "occasion": "Festive & Party Wear"},
            "amazon": {
                "title": "Kalarang Sarees Women's Georgette Floral Printed Saree with Unstitched Blouse Piece (Emerald Green, 5.5 m)",
                "bullet_points": [
                    "GEORGETTE DRAPE: Lightweight georgette fabric designed for effortless draping and pleating.",
                    "PALLU & BORDER DESIGN: Features detailed floral print motifs across the body and pallu.",
                    "COLORFAST & DURABLE: Follow the provided care label to help preserve the vibrancy of the print.",
                    "FESTIVE OCCASIONS: Suitable for family gatherings, festive puja, and evening celebrations.",
                    "DIMENSIONS: Saree length 5.5 meters with a separate 0.8 meter unstitched matching blouse piece."
                ],
                "backend_search_terms": "partywear ethnic attire traditional printed drape festival saree unstitched",
                "description": "Georgette floral printed saree by Kalarang Sarees with unstitched blouse piece for ethnic celebrations."
            },
            "flipkart": {
                "title": "Kalarang Sarees Floral Print Georgette Saree (Green)",
                "search_keywords": "georgette saree, printed saree women, festive saree",
                "description": "Floral printed georgette saree featuring matching unstitched blouse piece for traditional festive occasions.",
                "attributes": {"fabric": "Georgette", "pattern": "Floral Print", "saree_type": "Regular Saree", "occasion": "Festive", "blouse_piece": "Unstitched Blouse Piece", "saree_length": "5.5 m", "border_type": "Printed Border"}
            },
            "meesho": {
                "title": "Georgette Floral Print Saree with Blouse",
                "hinglish_hook_description": "Sundar floral printed georgette saree matching unstitched blouse piece ke saath. Festive aur party wear ke liye.",
                "english_hook_description": "Georgette floral printed saree with matching unstitched blouse piece for traditional festive occasions.",
                "highlights": ["Fabric: Georgette", "Length: 5.5m + 0.8m Blouse", "Work: Floral Print", "Care: Gentle Wash"]
            }
        },
        "coord_set": {
            "sku_id": "SKU_COORD_01",
            "category_profile": "coord_set",
            "brand": "Urban Stitch",
            "category": "Co-ord Sets",
            "product_type": "Pure Cotton 2-Piece Tunic and Trouser Set",
            "color": "Beige Rust",
            "sizes": "S, M, L, XL",
            "mrp": 2199.0,
            "meesho_price": 649.0,
            "seller_config": {"amazon_quantity": 35, "gst_percent": 5, "hsn_code": "62046200"},
            "verified": {"fabric": "Pure Cotton", "pattern": "Geometric Print", "top_type": "Tunic Top", "bottom_type": "Trousers", "sleeve": "3/4 Sleeve", "neckline": "Mandarin Neck", "top_length": "Hip Length (28 in)", "bottom_length": "Ankle Length (38 in)", "package_contents": "1 Top, 1 Bottom", "care_label": "Machine Wash Normal", "sizes": "S to XL", "color": "Beige Rust"},
            "amazon": {
                "title": "Urban Stitch Women's Pure Cotton 2-Piece Co-ord Set with Tunic Top and Trousers (Beige Rust)",
                "bullet_points": [
                    "100% PURE COTTON SET: Breathable woven cotton fabric tailored for coordinated 2-piece styling.",
                    "TOP & TROUSER SPECIFICATIONS: Mandarin neck hip-length tunic paired with elasticated ankle trousers.",
                    "COLORFAST & DURABLE: Follow the provided care label to help preserve the geometric print.",
                    "VERSATILE OCCASION WEAR: Designed for resort vacations, casual outings, or smart office wear.",
                    "SIZE & FIT GUIDE: True-to-size standard S (36 in) to XL (42 in) matching regular set dimensions."
                ],
                "backend_search_terms": "matching 2piece set vacation resortwear casual trouser tunic attire",
                "description": "Coordinated pure cotton 2-piece set by Urban Stitch comprising a geometric tunic top and ankle-length trousers."
            },
            "flipkart": {
                "title": "Urban Stitch Women Geometric Print Cotton Co-ords",
                "search_keywords": "cotton coord set, 2 piece suit women, printed pants top",
                "description": "Pure cotton 2-piece co-ord set featuring a mandarin collar tunic top and matching trousers for everyday wear.",
                "attributes": {"fabric": "Pure Cotton", "top_type": "Tunic Top", "bottom_type": "Trousers", "sleeve": "3/4 Sleeve", "neckline": "Mandarin Neck", "pattern": "Geometric Print", "occasion": "Casual", "package_contents": "1 Top, 1 Bottom"}
            },
            "meesho": {
                "title": "Pure Cotton 2-Piece Co-ord Set",
                "hinglish_hook_description": "Pure cotton 2-piece co-ord set. Tunic top aur matching trouser ka trendy combo daily wear aur travel ke liye.",
                "english_hook_description": "Pure cotton 2-piece co-ord set with geometric printed tunic top and comfortable trousers.",
                "highlights": ["Fabric: 100% Cotton", "Set: 1 Top + 1 Trouser", "Sizes: S to XL", "Care: Machine Wash"]
            }
        },
        "footwear": {
            "sku_id": "SKU_FOOTWEAR_01",
            "category_profile": "footwear",
            "brand": "StepCraft",
            "category": "Footwear",
            "product_type": "Men's Synthetic Leather Casual Loafers",
            "color": "Tan Brown",
            "sizes": "UK 6, UK 7, UK 8, UK 9, UK 10",
            "mrp": 1999.0,
            "meesho_price": 599.0,
            "seller_config": {"amazon_quantity": 45, "gst_percent": 12, "hsn_code": "64029990"},
            "verified": {"footwear_type": "Loafers", "material": "Synthetic Leather / PU", "closure_type": "Slip-On", "sole_material": "TPR (Thermoplastic Rubber)", "toe_shape": "Round Toe", "heel_type_or_flat": "Flat Sole", "care_label": "Wipe with a clean dry cloth", "sizes": "UK 6 to UK 10", "color": "Tan Brown"},
            "amazon": {
                "title": "StepCraft Men's Synthetic Leather Casual Slip-On Loafers with TPR Sole (Tan Brown)",
                "bullet_points": [
                    "SYNTHETIC LEATHER UPPER: Constructed with durable PU synthetic leather in classic tan finish.",
                    "TPR OUTSOLE CONSTRUCTION: Molded thermoplastic rubber sole designed for structured everyday wear.",
                    "MAINTENANCE & CARE: Wipe clean with a damp cloth; store away from moisture to maintain surface appearance.",
                    "VERSATILE PAIRING: Complements casual denim, chinos, and semi-formal trousers effortlessly.",
                    "STANDARD UK SIZING: Standard Indian / UK footwear sizing available from UK 6 to UK 10."
                ],
                "backend_search_terms": "mens shoes slipon moccasins driving shoes formal footwear flat shoes",
                "description": "Men's synthetic leather slip-on loafers by StepCraft with durable TPR sole and round toe construction."
            },
            "flipkart": {
                "title": "StepCraft Men Tan Slip-On Loafers",
                "search_keywords": "men loafers, casual shoes tan, slip on shoes",
                "description": "Tan brown synthetic leather casual loafers featuring slip-on closure and TPR sole for daily casual styling.",
                "attributes": {"footwear_type": "Loafers", "material": "Synthetic Leather / PU", "closure_type": "Slip-On", "sole_material": "TPR (Thermoplastic Rubber)", "toe_shape": "Round Toe", "heel_type": "Flat Sole", "occasion": "Casual", "size_system": "Indian / UK Sizing (e.g. UK 6 to UK 11)"}
            },
            "meesho": {
                "title": "Men's Tan Synthetic Leather Loafers",
                "hinglish_hook_description": "Tan brown synthetic leather casual loafers. Slip-on design aur durable TPR sole daily office aur party wear ke liye.",
                "english_hook_description": "Men's tan brown synthetic leather loafers with slip-on closure and sturdy TPR sole.",
                "highlights": ["Upper: Synthetic Leather", "Sole: TPR Sole", "Closure: Slip-On", "Sizes: UK 6 to UK 10"]
            }
        },
        "home_textiles": {
            "sku_id": "SKU_HOME_01",
            "category_profile": "home_textiles",
            "brand": "LoomNest",
            "category": "Home & Furnishing",
            "product_type": "100% Pure Cotton King Size Bedsheet with 2 Pillow Covers",
            "color": "Floral Blue",
            "sizes": "King Size (108 x 108 in / 274 x 274 cm)",
            "mrp": 1799.0,
            "meesho_price": 549.0,
            "seller_config": {"amazon_quantity": 60, "gst_percent": 12, "hsn_code": "63041910"},
            "verified": {"product_type": "Bedsheet Set", "material": "100% Pure Cotton", "pattern": "Floral Print", "dimensions": "King Size (108 x 108 in / 274 x 274 cm)", "package_contents": "1 Bedsheet with 2 Pillow Covers", "care_label": "Machine Wash Normal in Cold Water", "color": "Floral Blue"},
            "amazon": {
                "title": "LoomNest 100% Pure Cotton King Size Bedsheet with 2 Pillow Covers (108 x 108 in, Floral Blue)",
                "bullet_points": [
                    "100% PURE COTTON WEAVE: Woven from pure cotton yarns for standard household bedding.",
                    "ACCURATE DIMENSIONS: Large King Size sheet measuring 108 x 108 inches (274 x 274 cm) for full bed tuck-in.",
                    "COLORFAST & DURABLE: Follow the provided care label to help maintain color vibrancy through standard wash cycles.",
                    "HOME INTERIOR DECOR: Traditional floral motifs designed to complement modern and ethnic bedroom aesthetics.",
                    "PACKAGE CONTENTS & CARE: Includes 1 King Size bedsheet and 2 matching pillow covers (18 x 27 in). Machine washable."
                ],
                "backend_search_terms": "double bed sheet king size pillow covers cotton bedding mattress linen",
                "description": "100% pure cotton King Size floral bedsheet with 2 matching pillow covers by LoomNest for bedroom decor."
            },
            "flipkart": {
                "title": "LoomNest Floral King Cotton Bedsheet (Blue)",
                "search_keywords": "cotton bedsheet, king size bedsheet, bed cover pillow covers",
                "description": "Pure cotton King size bedsheet set with 2 pillow covers featuring floral print design.",
                "attributes": {"home_textile_type": "Bedsheet Set", "material": "100% Pure Cotton", "pattern": "Floral Print", "dimensions": "King Size (108 x 108 in / 274 x 274 cm)", "package_contents": "1 Bedsheet with 2 Pillow Covers", "care": "Machine Wash Normal"}
            },
            "meesho": {
                "title": "Pure Cotton King Size Bedsheet with 2 Covers",
                "hinglish_hook_description": "100% Pure Cotton King size bedsheet 2 pillow covers ke saath. Sundar floral print aur bada size 108x108 inch.",
                "english_hook_description": "100% pure cotton king size floral printed bedsheet with 2 matching pillow covers.",
                "highlights": ["Fabric: 100% Pure Cotton", "Size: King (108 x 108 in)", "Contents: 1 Sheet + 2 Covers", "Care: Machine Wash"]
            }
        }
    }

    if profile_id in samples:
        base_sample = samples[profile_id]
    else:
        # Generate generic valid sample for remaining profiles
        prof_name = prof["display_name"]
        cat_grp = prof["category_group"]
        req_facts = {f: "Standard Verified Specification" for f in prof["required_verified_fields"]}
        req_facts["care_label"] = "Standard Wash Care as per Label"
        req_facts["color"] = "Navy Blue"
        req_facts["sizes"] = "S, M, L, XL"
        req_facts["fabric"] = "Pure Cotton" if "fabric" in prof["required_verified_fields"] else "Standard Material"
        
        # Populate controlled attrs with valid first choice
        ctrl_attrs = {}
        for k, vlist in prof["controlled_attributes"].items():
            ctrl_attrs[k] = vlist[0]

        clean_name = prof['display_name'].split('(')[0].strip()
        base_sample = {
            "sku_id": f"SKU_{profile_id.upper()[:8]}_01",
            "category_profile": profile_id,
            "brand": "ProBrand",
            "category": cat_grp,
            "product_type": f"Standard {clean_name}",
            "color": "Navy Blue",
            "sizes": "S, M, L, XL",
            "mrp": 1499.0,
            "meesho_price": 449.0,
            "seller_config": {"amazon_quantity": 50, "gst_percent": 5, "hsn_code": "62114200"},
            "verified": req_facts,
            "amazon": {
                "title": f"ProBrand {clean_name} in Navy Blue",
                "bullet_points": [
                    "MATERIAL SPECIFICATION: Constructed with quality materials according to verified product records.",
                    "DESIGN & CRAFTSMANSHIP: Tailored construction built for regular everyday use.",
                    "COLORFAST & DURABLE: Follow the provided care label to help maintain the item's appearance.",
                    "VERSATILE STYLING: Pairs seamlessly across multiple casual and functional settings.",
                    "ACCURATE SIZING: Standard size range available from size S to XL."
                ],
                "backend_search_terms": "casual apparel daily attire quality top bottom clothing",
                "description": f"Standard {prof['display_name']} crafted by ProBrand for everyday styling."
            },
            "flipkart": {
                "title": f"ProBrand {prof['display_name'].split('(')[0].strip()} (Blue)",
                "search_keywords": f"{profile_id.replace('_', ' ')}, quality product",
                "description": f"Quality {prof['display_name']} designed for regular wear and versatile styling.",
                "attributes": ctrl_attrs
            },
            "meesho": {
                "title": f"ProBrand {prof['display_name'].split('(')[0].strip()[:35]}",
                "hinglish_hook_description": f"Daily wear aur regular use ke liye badhiya {prof['display_name'].split('(')[0].strip()}. Quality material ke saath.",
                "english_hook_description": f"Quality {prof['display_name'].split('(')[0].strip()} tailored for everyday versatile styling.",
                "highlights": ["Quality Material", "Regular Fit", "Multiple Sizes", "Standard Wash Care"]
            }
        }

    # Add standard 5x alternate variants
    alternates = []
    themes = [
        "Daily Office Workwear",
        "Festive & Celebration",
        "Summer Daily Wear",
        "Everyday Essential",
        "Modern Fusion Styling"
    ]
    for i, theme in enumerate(themes):
        alternates.append({
            "variant_id": f"V{i+1}",
            "angle_theme": theme,
            "amazon": {
                "title": f"{base_sample['brand']} {base_sample['product_type']} - {theme} Edition",
                "bullet_points": [
                    f"THEME SPECIFICATION: Tailored specifically for {theme.lower()} according to verified specifications.",
                    "DESIGN DETAILS: Functional construction suited for modern versatile lifestyles.",
                    "COLORFAST & DURABLE: Follow the provided care label to help preserve the fabric's appearance.",
                    "OCCASION WEAR: Designed to adapt across work, travel, and festive settings.",
                    "MAINTENANCE GUIDE: Follow the product care label for washing instructions."
                ],
                "backend_search_terms": f"apparel clothing {theme.lower().replace('&', '').replace('  ', ' ')} outfit attire",
                "description": f"{base_sample['product_type']} tailored specifically for {theme.lower()}."
            },
            "flipkart": {
                "title": f"{base_sample['brand']} {base_sample['product_type']} - {theme}",
                "search_keywords": f"{profile_id.replace('_', ' ')}, {theme.lower()}",
                "description": f"{base_sample['product_type']} designed for {theme.lower()} with verified attributes.",
                "attributes": base_sample["flipkart"]["attributes"]
            },
            "meesho": {
                "title": f"{base_sample['product_type'][:35]} - {theme[:15]}",
                "hinglish_hook_description": f"Har din ke liye badhiya! {theme} ke liye ready look aur quality material.",
                "english_hook_description": f"Crafted specifically for {theme.lower()} and versatile styling.",
                "highlights": ["Verified Quality", f"Theme: {theme[:18]}", "Standard Fit", "Wash as per Label"]
            }
        })
    base_sample["alternates"] = alternates
    return [base_sample]


def get_sample_skus() -> list[dict[str, Any]]:
    """Default multi-profile sample SKU batch for studio initial load."""
    return get_sample_skus_for_profile("women_ethnic_kurta")


def normalize_and_merge_json(raw_data: Any) -> tuple[Optional[str], Optional[BatchConfig], list[dict]]:
    schema_version = None
    batch_cfg = None
    skus_raw = []

    if isinstance(raw_data, dict):
        schema_version = raw_data.get("schema_version")
        if "batch_config" in raw_data and isinstance(raw_data["batch_config"], dict):
            batch_cfg = BatchConfig(**raw_data["batch_config"])

        if "skus" in raw_data and isinstance(raw_data["skus"], list):
            skus_raw = raw_data["skus"]
        elif "sku_id" in raw_data:
            skus_raw = [raw_data]
    elif isinstance(raw_data, list):
        skus_raw = raw_data

    merged = []
    for item in skus_raw:
        if isinstance(item, dict):
            norm_item, _ = normalize_sku_profile_and_attributes(item, batch_cfg)
            merged.append(norm_item)

    return schema_version, batch_cfg, merged


# ──────────────────────────────────────────────
# FastAPI Application
# ──────────────────────────────────────────────

app = FastAPI(title="Listing Factory", version="2.1.0")


@app.get("/api/category-profiles")
async def get_category_profiles():
    """Returns the central registry of supported category profiles and exclusions."""
    return JSONResponse({
        "supported_profiles": CATEGORY_PROFILES,
        "excluded_profiles": EXCLUDED_PROFILES
    })


@app.post("/api/validate-json")
async def validate_json(request: Request):
    """
    Strictly validate AI-generated JSON listing data against Category Profile specifications.
    """
    try:
        body = await request.json()
        raw = body.get("json_text", "")
        if not raw or not raw.strip():
            return JSONResponse({
                "valid": False,
                "global_errors": ["JSON payload is empty."],
                "global_warnings": [],
                "sku_results": []
            }, status_code=400)
        
        data = json.loads(raw)
        schema_ver, batch_cfg, merged_skus = normalize_and_merge_json(data)

        global_warnings = []
        global_errors = []

        if schema_ver is not None and schema_ver != EXPECTED_SCHEMA_VERSION:
            global_errors.append(
                f"Expected schema_version '{EXPECTED_SCHEMA_VERSION}', got '{schema_ver}'. "
                "Please regenerate JSON using the current prompt."
            )
            return JSONResponse({
                "valid": False,
                "schema_version": schema_ver,
                "expected_schema_version": EXPECTED_SCHEMA_VERSION,
                "global_errors": global_errors,
                "global_warnings": [],
                "sku_results": []
            }, status_code=400)

        if len(merged_skus) > MAX_SKUS_SOFT_LIMIT:
            global_warnings.append(
                f"Batch has {len(merged_skus)} SKUs (recommended soft limit: {MAX_SKUS_SOFT_LIMIT}). "
                "Consider splitting into smaller batches for optimal performance."
            )

        sku_results = []
        parsed_skus = []
        overall_valid = True

        for idx, item in enumerate(merged_skus):
            sk_id = item.get("sku_id") or f"SKU_{idx+1:02d}"
            sku_errors = []
            sku_warnings = []

            # Check for "To be confirmed" fields
            for key, val in item.items():
                if isinstance(val, str) and val.strip().lower() == TO_BE_CONFIRMED.lower():
                    sku_warnings.append(f"Field '{key}' is marked '{TO_BE_CONFIRMED}'")
                elif isinstance(val, dict):
                    for subk, subv in val.items():
                        if isinstance(subv, str) and subv.strip().lower() == TO_BE_CONFIRMED.lower():
                            sku_warnings.append(f"{key}.{subk} is marked '{TO_BE_CONFIRMED}'")

            prof_id = item.get("category_profile", "women_ethnic_kurta")
            if prof_id in EXCLUDED_PROFILES:
                sku_errors.append(f"Category profile '{prof_id}' ({EXCLUDED_PROFILES[prof_id]}) is intentionally not supported in Listing Factory v2.0.")

            try:
                sku_obj = SKUItem(**item)
                parsed_skus.append(sku_obj)

                # Category Profile Validation
                prof_errs, prof_warns = validate_sku_category_profile(sku_obj)
                sku_errors.extend(prof_errs)
                sku_warnings.extend(prof_warns)

                # Near-limit advisory checks
                t_len = len(sku_obj.amazon.title)
                if sku_obj.amazon.title != TO_BE_CONFIRMED and t_len > 165:
                    sku_warnings.append(f"Amazon title length is near limit ({t_len}/180 chars)")
                
                b_len = len(sku_obj.amazon.backend_search_terms.encode("utf-8"))
                if sku_obj.amazon.backend_search_terms != TO_BE_CONFIRMED and b_len > 225:
                    sku_warnings.append(f"Amazon search terms bytes near limit ({b_len}/240 bytes)")
                
                m_len = len(sku_obj.meesho.title)
                if sku_obj.meesho.title != TO_BE_CONFIRMED and m_len > 55:
                    sku_warnings.append(f"Meesho title length is near limit ({m_len}/60 chars)")

                # Amazon Title Quality
                atq_errs, atq_warns = validate_amazon_title_quality(sku_obj.amazon.title, sku_obj.brand)
                sku_errors.extend(atq_errs)
                sku_warnings.extend(atq_warns)

                # BST Hygiene
                bst_errs, bst_warns = validate_amazon_backend_search_terms(
                    sku_obj.amazon.title, sku_obj.amazon.backend_search_terms, sku_obj.brand
                )
                sku_errors.extend(bst_errs)
                sku_warnings.extend(bst_warns)

                # Bullet 3 Safety
                if len(sku_obj.amazon.bullet_points) >= 3:
                    b3_errs, b3_warns = validate_bullet_3_safety(sku_obj.amazon.bullet_points[2])
                    sku_errors.extend(b3_errs)
                    sku_warnings.extend(b3_warns)

                # Truthfulness & Claim Safety
                fields_to_check = [
                    ("amazon.title", sku_obj.amazon.title),
                    ("amazon.description", sku_obj.amazon.description),
                    ("flipkart.title", sku_obj.flipkart.title),
                    ("flipkart.search_keywords", sku_obj.flipkart.search_keywords),
                    ("flipkart.description", sku_obj.flipkart.description),
                    ("meesho.title", sku_obj.meesho.title),
                    ("meesho.hinglish_hook_description", sku_obj.meesho.hinglish_hook_description),
                    ("meesho.english_hook_description", sku_obj.meesho.english_hook_description),
                ]
                for bi, bp in enumerate(sku_obj.amazon.bullet_points):
                    fields_to_check.append((f"amazon.bullet_points[{bi+1}]", bp))
                for hi, hl in enumerate(sku_obj.meesho.highlights):
                    fields_to_check.append((f"meesho.highlights[{hi+1}]", hl))

                for fpath, ftext in fields_to_check:
                    c_errs, c_warns = validate_truthfulness_and_claims(fpath, ftext, sku_obj.category_profile, sku_obj.verified)
                    sku_errors.extend(c_errs)
                    sku_warnings.extend(c_warns)

                # Validate alternates
                for alt in sku_obj.alternates:
                    alt_atq_errs, alt_atq_warns = validate_amazon_title_quality(alt.amazon.title, sku_obj.brand)
                    sku_errors.extend([f"Alternate {alt.variant_id}: {e}" for e in alt_atq_errs])
                    sku_warnings.extend([f"Alternate {alt.variant_id}: {w}" for w in alt_atq_warns])

                    alt_bst_errs, alt_bst_warns = validate_amazon_backend_search_terms(
                        alt.amazon.title, alt.amazon.backend_search_terms, sku_obj.brand
                    )
                    sku_errors.extend([f"Alternate {alt.variant_id}: {e}" for e in alt_bst_errs])
                    sku_warnings.extend([f"Alternate {alt.variant_id}: {w}" for w in alt_bst_warns])

                    if len(alt.amazon.bullet_points) >= 3:
                        alt_b3_errs, alt_b3_warns = validate_bullet_3_safety(alt.amazon.bullet_points[2])
                        sku_errors.extend([f"Alternate {alt.variant_id}: {e}" for e in alt_b3_errs])

                    alt_fields = [
                        (f"alternates.{alt.variant_id}.amazon.title", alt.amazon.title),
                        (f"alternates.{alt.variant_id}.amazon.description", alt.amazon.description),
                        (f"alternates.{alt.variant_id}.flipkart.title", alt.flipkart.title),
                        (f"alternates.{alt.variant_id}.flipkart.search_keywords", alt.flipkart.search_keywords),
                        (f"alternates.{alt.variant_id}.flipkart.description", alt.flipkart.description),
                        (f"alternates.{alt.variant_id}.meesho.title", alt.meesho.title),
                        (f"alternates.{alt.variant_id}.meesho.hinglish_hook_description", alt.meesho.hinglish_hook_description),
                        (f"alternates.{alt.variant_id}.meesho.english_hook_description", alt.meesho.english_hook_description),
                    ]
                    for bi, bp in enumerate(alt.amazon.bullet_points):
                        alt_fields.append((f"alternates.{alt.variant_id}.amazon.bullet_points[{bi+1}]", bp))
                    for hi, hl in enumerate(alt.meesho.highlights):
                        alt_fields.append((f"alternates.{alt.variant_id}.meesho.highlights[{hi+1}]", hl))

                    for fpath, ftext in alt_fields:
                        alt_c_errs, alt_c_warns = validate_truthfulness_and_claims(fpath, ftext, sku_obj.category_profile, sku_obj.verified)
                        sku_errors.extend(alt_c_errs)
                        sku_warnings.extend(alt_c_warns)

            except Exception as pe:
                sku_errors.append(str(pe))

            if sku_errors:
                overall_valid = False

            sku_results.append({
                "sku_id": sk_id,
                "category_profile": prof_id,
                "valid": len(sku_errors) == 0,
                "errors": sku_errors,
                "warnings": sku_warnings,
            })

        if not overall_valid:
            return JSONResponse({
                "valid": False,
                "schema_version": schema_ver or EXPECTED_SCHEMA_VERSION,
                "expected_schema_version": EXPECTED_SCHEMA_VERSION,
                "sku_count": len(merged_skus),
                "global_errors": global_errors,
                "global_warnings": global_warnings,
                "sku_results": sku_results,
            }, status_code=400)

        return JSONResponse({
            "valid": True,
            "schema_version": schema_ver or EXPECTED_SCHEMA_VERSION,
            "expected_schema_version": EXPECTED_SCHEMA_VERSION,
            "sku_count": len(parsed_skus),
            "global_errors": [],
            "global_warnings": global_warnings,
            "sku_results": sku_results,
        })

    except json.JSONDecodeError as e:
        return JSONResponse({
            "valid": False,
            "global_errors": [f"Invalid JSON syntax at line {e.lineno}, column {e.colno}: {e.msg}"],
            "global_warnings": [],
            "sku_results": []
        }, status_code=400)
    except Exception as e:
        return JSONResponse({
            "valid": False,
            "global_errors": [str(e)],
            "global_warnings": [],
            "sku_results": []
        }, status_code=400)


@app.post("/api/generate")
async def generate_package(
    client_name: str = Form(...),
    batch_id: str = Form(...),
    category: str = Form(...),
    json_data: str = Form(...),
    images: list[UploadFile] = File(default=[]),
):
    try:
        raw = json.loads(json_data)
        schema_ver, batch_cfg, merged_skus = normalize_and_merge_json(raw)
        skus = [SKUItem(**item) for item in merged_skus]
    except Exception as e:
        return JSONResponse({"error": f"Schema validation error: {e}"}, status_code=400)

    image_files: list[tuple[str, bytes]] = []
    total_bytes = 0
    for img in images:
        data = await img.read()
        total_bytes += len(data)
        image_files.append((img.filename, data))

    safe_client = re.sub(r"[^\w\-]", "_", client_name.strip() or "Client")
    safe_batch = re.sub(r"[^\w\-]", "_", batch_id.strip() or "Batch_01")
    cat = category.strip() or "Women Ethnic Wear"
    ver_str = schema_ver or EXPECTED_SCHEMA_VERSION

    zip_bytes = build_zip(
        safe_client, safe_batch, cat, skus, image_files,
        validation_status="✅ Pass", schema_version=ver_str
    )

    pattern = str(OUTPUT_DIR / f"{safe_client}_{safe_batch}_v*.zip")
    existing_files = glob.glob(pattern)
    existing_versions = []
    for f in existing_files:
        m = re.search(r"_v(\d+)\.zip$", f)
        if m:
            existing_versions.append(int(m.group(1)))
    next_ver = (max(existing_versions) + 1) if existing_versions else 1

    out_filename = f"{safe_client}_{safe_batch}_v{next_ver}.zip"
    out_path = OUTPUT_DIR / out_filename
    with open(out_path, "wb") as f:
        f.write(zip_bytes)

    headers = {
        "Content-Disposition": f'attachment; filename="{out_filename}"',
        "Content-Type": "application/zip",
        "X-Archive-Version": f"v{next_ver}",
    }
    return StreamingResponse(io.BytesIO(zip_bytes), headers=headers, media_type="application/zip")


@app.post("/api/generate-sample")
async def generate_sample(profile_id: Optional[str] = None):
    prof_id = profile_id or "women_ethnic_kurta"
    skus_raw = get_sample_skus_for_profile(prof_id)
    skus = [SKUItem(**d) for d in skus_raw]

    dummy_images: list[tuple[str, bytes]] = []
    for sku in skus:
        for sfx in ["_MAIN", "_PT01", "_PT02", "_PT03", "_PT04", "_PT05"]:
            fname = f"{sku.sku_id}{sfx}.jpg"
            dummy_images.append((fname, b"\xFF\xD8\xFF\xE0\x00\x10JFIF" + b"\x00" * 64))

    zip_bytes = build_zip(
        client_name="SampleClient",
        batch_id="SampleBatch",
        category="Women Ethnic Wear",
        skus=skus,
        image_files=dummy_images,
        validation_status="✅ Pass",
        schema_version=EXPECTED_SCHEMA_VERSION,
    )

    headers = {
        "Content-Disposition": f'attachment; filename="SampleClient_SampleBatch_{prof_id}_v1.zip"',
        "Content-Type": "application/zip",
    }
    return StreamingResponse(io.BytesIO(zip_bytes), headers=headers, media_type="application/zip")


@app.get("/api/history")
async def list_history():
    files = sorted(OUTPUT_DIR.glob("*.zip"), key=os.path.getmtime, reverse=True)
    results = []
    for f in files:
        stat = f.stat()
        m = re.search(r"^(.*?)_(.*?)_v(\d+)\.zip$", f.name)
        client = m.group(1) if m else "Unknown"
        batch = m.group(2) if m else "Unknown"
        ver = f"v{m.group(3)}" if m else "v1"
        results.append({
            "filename": f.name,
            "client_name": client,
            "batch_id": batch,
            "version": ver,
            "size_kb": round(stat.st_size / 1024, 1),
            "created_at": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        })
    return JSONResponse({"packages": results})


@app.get("/", response_class=HTMLResponse)
async def serve_index():
    index_file = Path(__file__).parent / "index.html"
    if index_file.exists():
        with open(index_file, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse("<h1>Listing Factory v2.1</h1><p>index.html not found.</p>")


if __name__ == "__main__":
    print("\n  [*] Listing Factory v2.1 - Category-Profile-Driven Studio")
    print("  ===============================================================")
    print("  > Server running at: http://127.0.0.1:8000")
    print("  > Supported Profiles: 13 distinct product families")
    print("  > Press Ctrl+C to stop\n")
    uvicorn.run("app:app", host="127.0.0.1", port=8000, reload=False)
