"""
Comprehensive Test Suite for Listing Factory v2.1 Category Profile Architecture
================================================================================
Tests:
1. 13 Supported Profiles Positive Validation
2. Profile-Specific Required Facts Validation (Negative Tests)
3. Excluded Profiles Rejection (Blouse, Lingerie, Innerwear, Shapewear, etc.)
4. Profile-Specific Claim Safety Tests (Footwear, Saree, Home Textiles, Kidswear, Shirts)
5. Backward Compatibility (Legacy Women Ethnic Wear inference)
6. Excel Builder Profile Dynamism & Column Integrity
7. Full Package ZIP Generation & Cryptographic Metadata
"""

import io
import json
import os
import sys
import zipfile
from pathlib import Path

# Add project root to sys.path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from openpyxl import load_workbook

from app import (
    CATEGORY_PROFILES,
    EXCLUDED_PROFILES,
    EXPECTED_SCHEMA_VERSION,
    STRUCTURAL_READINESS_DISCLAIMER,
    IMAGE_ROLE_DISCLAIMER,
    SKUItem,
    get_sample_skus_for_profile,
    normalize_and_merge_json,
    validate_sku_category_profile,
    validate_amazon_title_quality,
    validate_truthfulness_and_claims,
    validate_amazon_backend_search_terms,
    build_workbook,
    build_alternates_workbook,
    build_package_metadata,
    generate_readme,
    build_zip,
)

def run_tests():
    passed = 0
    failed = 0

    def assert_true(condition, msg):
        nonlocal passed, failed
        if condition:
            passed += 1
            print(f"  [PASS] {msg}")
        else:
            failed += 1
            print(f"  [FAIL] {msg}")

    print("\n=======================================================")
    print("  TEST SUITE 1: Positive Validation for All 13 Profiles")
    print("=======================================================")

    all_13_profiles = list(CATEGORY_PROFILES.keys())
    assert_true(len(all_13_profiles) == 13, f"Registry contains exactly 13 supported profiles (got {len(all_13_profiles)})")

    for prof_id in all_13_profiles:
        sample_list = get_sample_skus_for_profile(prof_id)
        assert_true(len(sample_list) == 1, f"Sample generated for profile '{prof_id}'")
        sku_data = sample_list[0]
        
        # Test Pydantic model parsing
        sku_obj = SKUItem(**sku_data)
        assert_true(sku_obj.category_profile == prof_id, f"SKUItem parsed with profile '{prof_id}'")
        
        # Test Profile Validation
        errs, warns = validate_sku_category_profile(sku_obj)
        assert_true(len(errs) == 0, f"Profile '{prof_id}' validation has 0 errors (got: {errs})")
        
        # Test Title Quality & Deduplication
        atq_errs, _ = validate_amazon_title_quality(sku_obj.amazon.title, sku_obj.brand)
        assert_true(len(atq_errs) == 0, f"Profile '{prof_id}' title quality has 0 errors (got: {atq_errs})")
        
        # Test Claims Safety
        c_errs, _ = validate_truthfulness_and_claims("amazon.title", sku_obj.amazon.title, prof_id, sku_obj.verified)
        assert_true(len(c_errs) == 0, f"Profile '{prof_id}' title claims have 0 errors (got: {c_errs})")

    print("\n=======================================================")
    print("  TEST SUITE 2: Negative Fact Requirement Validation")
    print("=======================================================")

    # 1. Saree missing saree_length
    saree_sample = get_sample_skus_for_profile("saree")[0]
    saree_bad = json.loads(json.dumps(saree_sample))
    del saree_bad["verified"]["saree_length"]
    if "saree_length" in saree_bad["flipkart"]["attributes"]:
        del saree_bad["flipkart"]["attributes"]["saree_length"]
    s_obj = SKUItem(**saree_bad)
    errs, _ = validate_sku_category_profile(s_obj)
    assert_true(any("saree_length" in e for e in errs), "Saree missing saree_length failed validation correctly")

    # 2. Co-ord set missing package_contents
    coord_sample = get_sample_skus_for_profile("coord_set")[0]
    coord_bad = json.loads(json.dumps(coord_sample))
    del coord_bad["verified"]["package_contents"]
    if "package_contents" in coord_bad["flipkart"]["attributes"]:
        del coord_bad["flipkart"]["attributes"]["package_contents"]
    c_obj = SKUItem(**coord_bad)
    errs, _ = validate_sku_category_profile(c_obj)
    assert_true(any("package_contents" in e for e in errs), "Co-ord set missing package_contents failed validation correctly")

    # 3. Footwear missing sole_material
    fw_sample = get_sample_skus_for_profile("footwear")[0]
    fw_bad = json.loads(json.dumps(fw_sample))
    del fw_bad["verified"]["sole_material"]
    if "sole_material" in fw_bad["flipkart"]["attributes"]:
        del fw_bad["flipkart"]["attributes"]["sole_material"]
    fw_obj = SKUItem(**fw_bad)
    errs, _ = validate_sku_category_profile(fw_obj)
    assert_true(any("sole_material" in e for e in errs), "Footwear missing sole_material failed validation correctly")

    # 4. Home textile missing dimensions
    ht_sample = get_sample_skus_for_profile("home_textiles")[0]
    ht_bad = json.loads(json.dumps(ht_sample))
    del ht_bad["verified"]["dimensions"]
    if "dimensions" in ht_bad["flipkart"]["attributes"]:
        del ht_bad["flipkart"]["attributes"]["dimensions"]
    ht_obj = SKUItem(**ht_bad)
    errs, _ = validate_sku_category_profile(ht_obj)
    assert_true(any("dimensions" in e for e in errs), "Home textiles missing dimensions failed validation correctly")

    # 5. Kidswear missing age_group
    kids_sample = get_sample_skus_for_profile("kidswear")[0]
    kids_bad = json.loads(json.dumps(kids_sample))
    del kids_bad["verified"]["age_group"]
    if "age_group" in kids_bad["flipkart"]["attributes"]:
        del kids_bad["flipkart"]["attributes"]["age_group"]
    kids_obj = SKUItem(**kids_bad)
    errs, _ = validate_sku_category_profile(kids_obj)
    assert_true(any("age_group" in e for e in errs), "Kidswear missing age_group failed validation correctly")

    print("\n=======================================================")
    print("  TEST SUITE 3: Excluded Profiles Rejection")
    print("=======================================================")

    for exc_key in ["blouse", "lingerie", "innerwear", "shapewear", "bra", "underwear"]:
        dummy_excluded = {
            "sku_id": "SKU_EXC_01",
            "category_profile": exc_key,
            "brand": "TestBrand",
            "product_type": "Test Item",
            "color": "Red",
            "sizes": "Free Size",
            "mrp": 999.0,
            "meesho_price": 299.0,
            "seller_config": {"amazon_quantity": 50, "gst_percent": 5, "hsn_code": "62114200"},
            "amazon": {
                "title": "TestBrand Women Red Item",
                "bullet_points": ["BP1", "BP2", "COLORFAST & DURABLE: Care label.", "BP4", "BP5"],
                "backend_search_terms": "test search terms",
                "description": "Test description"
            },
            "flipkart": {"title": "TestBrand Item", "search_keywords": "test", "description": "test"},
            "meesho": {"title": "Test Item", "hinglish_hook_description": "test", "english_hook_description": "test", "highlights": ["H1", "H2", "H3", "H4"]}
        }
        sku_exc_obj = SKUItem(**dummy_excluded)
        errs, _ = validate_sku_category_profile(sku_exc_obj)
        assert_true(
            any("intentionally not supported in Listing Factory" in e for e in errs),
            f"Excluded profile '{exc_key}' rejected with intentional exclusion error"
        )

    print("\n=======================================================")
    print("  TEST SUITE 4: Profile-Specific Claim Boundary Tests")
    print("=======================================================")

    # Footwear claiming unverified anti-slip
    c_errs, _ = validate_truthfulness_and_claims("amazon.title", "Men's anti-slip loafers", "footwear", {})
    assert_true(any("anti-slip" in e for e in c_errs), "Unverified footwear anti-slip claim blocked")

    # Footwear claiming unverified cushioning
    c_errs, _ = validate_truthfulness_and_claims("amazon.description", "Features memory foam cushioning", "footwear", {})
    assert_true(any("cushioning" in e or "memory foam" in e for e in c_errs), "Unverified footwear cushioning claim blocked")

    # Saree claiming unverified handloom
    c_errs, _ = validate_truthfulness_and_claims("amazon.title", "Traditional Handloom Saree", "saree", {})
    assert_true(any("handloom" in e for e in c_errs), "Unverified saree handloom claim blocked")

    # Saree claiming unverified Banarasi
    c_errs, _ = validate_truthfulness_and_claims("flipkart.title", "Women Banarasi Silk Saree", "saree", {})
    assert_true(any("Banarasi" in e for e in c_errs), "Unverified saree Banarasi claim blocked")

    # Home textile claiming unverified thread count
    c_errs, _ = validate_truthfulness_and_claims("amazon.title", "400 TC Pure Cotton Bedsheet", "home_textiles", {})
    assert_true(any("thread count" in e for e in c_errs), "Unverified home textile thread count claim blocked")

    # Kidswear claiming unverified skin-safe
    c_errs, _ = validate_truthfulness_and_claims("meesho.hinglish_hook_description", "Skin-safe cotton frock for girls", "kidswear", {})
    assert_true(any("skin-safe" in e for e in c_errs), "Unverified kidswear skin-safe claim blocked")

    # Shirt unverified slim fit warning
    _, c_warns = validate_truthfulness_and_claims("amazon.title", "Men's Slim Fit Shirt", "men_shirt", {})
    assert_true(any("slim fit" in w for w in c_warns), "Unverified slim fit descriptor triggers advisory warning")

    print("\n=======================================================")
    print("  TEST SUITE 5: Legacy v2.0 Backward Compatibility")
    print("=======================================================")

    legacy_payload = {
        "schema_version": "v2.0",
        "batch_config": {
            "brand": "Anvi Fabrics",
            "category": "Women Ethnic Wear",
            "seller_config": {"amazon_quantity": 50, "gst_percent": 5, "hsn_code": "62114200"}
        },
        "skus": [{
            "sku_id": "SKU_LEGACY_01",
            "product_type": "Pure Cotton Anarkali Kurti",
            "color": "Navy Blue",
            "sizes": "S, M, L, XL",
            "mrp": 1299,
            "meesho_price": 349,
            "amazon": {
                "title": "Anvi Fabrics Women's Pure Cotton Anarkali Kurti in Navy Blue",
                "bullet_points": ["BP1", "BP2", "COLORFAST & DURABLE: Care label.", "BP4", "BP5"],
                "backend_search_terms": "tunic top printed attire",
                "description": "Pure cotton anarkali kurti"
            },
            "flipkart": {
                "title": "Anvi Fabrics Printed Cotton Kurti",
                "fabric": "Pure Cotton",
                "kurta_type": "Anarkali",
                "neck": "Mandarin Neck",
                "sleeve": "3/4 Sleeve",
                "length_type": "Calf Length",
                "pattern": "Floral Print",
                "occasion": "Casual",
                "search_keywords": "kurti",
                "description": "Pure cotton anarkali"
            },
            "meesho": {
                "title": "Cotton Anarkali Kurti",
                "hinglish_hook_description": "Daily wear kurti",
                "english_hook_description": "Daily wear kurti",
                "highlights": ["Fabric: Cotton", "Style: Anarkali", "Sizes: S-XL", "Care: Wash"]
            }
        }]
    }

    schema_ver, batch_cfg, merged = normalize_and_merge_json(legacy_payload)
    assert_true(schema_ver == "v2.0", "Legacy schema_version preserved")
    assert_true(len(merged) == 1, "Legacy SKU parsed")
    assert_true(merged[0]["category_profile"] == "women_ethnic_kurta", "Legacy Women Ethnic Wear inferred as 'women_ethnic_kurta'")
    assert_true(merged[0]["flipkart"]["attributes"]["kurta_type"] == "Anarkali", "Legacy flat Flipkart kurta_type migrated into attributes")

    print("\n=======================================================")
    print("  TEST SUITE 6: Profile-Aware Excel Workbook Integrity")
    print("=======================================================")

    # Test multi-profile workbook generation
    saree_sku = SKUItem(**get_sample_skus_for_profile("saree")[0])
    footwear_sku = SKUItem(**get_sample_skus_for_profile("footwear")[0])
    home_sku = SKUItem(**get_sample_skus_for_profile("home_textiles")[0])
    test_skus = [saree_sku, footwear_sku, home_sku]

    img_counts = {s.sku_id: [f"{s.sku_id}_MAIN.jpg", f"{s.sku_id}_PT01.jpg", f"{s.sku_id}_PT02.jpg", f"{s.sku_id}_PT03.jpg"] for s in test_skus}
    wb_bytes = build_workbook(test_skus, "Mixed", img_counts, "✅ Pass")

    wb = load_workbook(io.BytesIO(wb_bytes))
    assert_true("Master_Summary" in wb.sheetnames, "Master_Summary sheet present")
    assert_true("01_Amazon_Bulk_Import" in wb.sheetnames, "01_Amazon_Bulk_Import sheet present")
    assert_true("02_Flipkart_Bulk_Import" in wb.sheetnames, "02_Flipkart_Bulk_Import sheet present")
    assert_true("03_Meesho_Bulk_Import" in wb.sheetnames, "03_Meesho_Bulk_Import sheet present")

    # Check Master_Summary 17-column structure
    ws_sum = wb["Master_Summary"]
    headers_sum = [cell.value for cell in ws_sum[1]]
    assert_true(len(headers_sum) == 17, f"Master_Summary has 17 columns (got {len(headers_sum)})")
    assert_true("Category Profile" in headers_sum, "Category Profile column present in Master_Summary")
    assert_true("Status Scope / Meaning" in headers_sum, "Status Scope column present in Master_Summary")

    # Check Flipkart dynamic columns
    ws_fk = wb["02_Flipkart_Bulk_Import"]
    headers_fk = [cell.value for cell in ws_fk[1]]
    assert_true("Category Profile" in headers_fk, "Category Profile column present in Flipkart sheet")
    assert_true("Sole Material" in headers_fk or "Saree Length" in headers_fk, "Dynamic profile attribute columns created in Flipkart sheet")

    # Test Alternates Workbook
    alt_bytes = build_alternates_workbook(test_skus)
    wb_alt = load_workbook(io.BytesIO(alt_bytes))
    assert_true("Amazon_Alternate_Copies" in wb_alt.sheetnames, "Amazon_Alternate_Copies tab present")
    assert_true("Flipkart_Alternate_Copies" in wb_alt.sheetnames, "Flipkart_Alternate_Copies tab present")
    assert_true("Meesho_Alternate_Copies" in wb_alt.sheetnames, "Meesho_Alternate_Copies tab present")

    print("\n=======================================================")
    print("  TEST SUITE 7: ZIP Package & Metadata Verification")
    print("=======================================================")

    dummy_images = [
        ("SKU_SAREE_01_MAIN.jpg", b"fake_img"),
        ("SKU_SAREE_01_PT01.jpg", b"fake_img"),
        ("SKU_FOOTWEAR_01_MAIN.jpg", b"fake_img"),
        ("Unassigned_Asset_01.jpg", b"fake_img")
    ]
    zip_bytes = build_zip("TestClient", "Batch_99", "Mixed", test_skus, dummy_images, "✅ Pass", EXPECTED_SCHEMA_VERSION)
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    file_list = zf.namelist()

    prefix = "TestClient_Batch_99"
    assert_true(f"{prefix}/{prefix}_Master_Marketplace_Upload.xlsx" in file_list, "Master Excel in ZIP")
    assert_true(f"{prefix}/{prefix}_Alternate_Listing_Copies.xlsx" in file_list, "Alternates Excel in ZIP")
    assert_true(f"{prefix}/{prefix}_package_metadata.json" in file_list, "Audit Metadata JSON in ZIP")
    assert_true(f"{prefix}/README.txt" in file_list, "README.txt in ZIP")
    assert_true(f"{prefix}/Organized_SKU_Images/SKU_SAREE_01/SKU_SAREE_01_MAIN.jpg" in file_list, "Saree hero image in SKU folder")
    assert_true(f"{prefix}/Organized_SKU_Images/Unassigned_Assets/Unassigned_Asset_01.jpg" in file_list, "Unassigned image in Unassigned folder")

    # Verify metadata JSON contents
    meta_raw = zf.read(f"{prefix}/{prefix}_package_metadata.json").decode("utf-8")
    meta_obj = json.loads(meta_raw)
    assert_true(meta_obj["tool_version"] == "Listing Factory v2.1", "Tool version v2.1 in metadata")
    assert_true(len(meta_obj["skus"]) == 3, "All 3 SKUs in metadata")
    assert_true(meta_obj["skus"][0]["category_profile"] == "saree", "Saree category profile recorded in metadata")

    print("\n=======================================================")
    print(f"  TOTAL RESULTS: {passed} PASSED, {failed} FAILED")
    print("=======================================================\n")
    return failed == 0

if __name__ == "__main__":
    success = run_tests()
    if not success:
        exit(1)
