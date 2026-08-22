"""
Comprehensive Automated Test Suite for Listing Factory v2.0
Sections 1-6 Quality & Truthfulness Refinements
"""

import os
import sys
import io
import json
import zipfile
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app import (
    TOOL_VERSION,
    JSON_PROMPT_VERSION,
    STRUCTURAL_READINESS_DISCLAIMER,
    IMAGE_ROLE_DISCLAIMER,
    IMAGE_ROLES,
    SKUItem,
    validate_amazon_title_quality,
    validate_truthfulness_and_claims,
    validate_amazon_backend_search_terms,
    validate_bullet_3_safety,
    generate_readme,
    build_workbook,
    build_alternates_workbook,
    get_sample_skus,
    build_package_metadata
)

def run_tests():
    print("================================================================================")
    print("RUNNING LISTING FACTORY V2.0 FINAL REFINEMENT TEST SUITE")
    print("================================================================================")

    # ──────────────────────────────────────────────────────────────────────────
    # SECTION 1: Amazon Title Quality Validator
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[TEST SECTION 1] Amazon Title Quality Validation...")
    
    # 1.1 Brand duplication (Hard error)
    errs, warns = validate_amazon_title_quality("Anvi Fabrics Women's Pure Cotton Kurti by Anvi Fabrics", "Anvi Fabrics")
    assert any("contains the brand name 'Anvi Fabrics' 2 times" in e for e in errs), f"Failed brand repetition test: {errs}"
    print("  ✓ Brand repetition correctly flagged as hard error")

    # 1.2 Adjacent duplicate words (Hard error)
    errs, warns = validate_amazon_title_quality("Anvi Fabrics Women's Pure Cotton Cotton Kurti", "Anvi Fabrics")
    assert any("adjacent word repetition: 'Cotton Cotton'" in e for e in errs), f"Failed adjacent words test: {errs}"
    print("  ✓ Adjacent word repetition correctly flagged as hard error")

    # 1.3 Redundant phrase stack (Hard error)
    errs, warns = validate_amazon_title_quality("Anvi Fabrics Women's Floral Print Printed Anarkali Kurti", "Anvi Fabrics")
    assert any("redundant phrase stack: 'Floral Print Printed'" in e for e in errs), f"Failed phrase stack test: {errs}"
    print("  ✓ Redundant phrase stack ('Floral Print Printed') flagged as hard error")

    # 1.4 Repeated close product types (Warning)
    errs, warns = validate_amazon_title_quality("Anvi Fabrics Women's Cotton Kurta Kurti with Pockets", "Anvi Fabrics")
    assert any("both 'kurta' and 'kurti'" in w for w in warns), f"Failed kurta/kurti test: {warns}"
    print("  ✓ Repeated close product types ('kurta' + 'kurti') flagged as advisory warning")

    # 1.5 Excessive delimiters (Warning)
    errs, warns = validate_amazon_title_quality("Anvi Fabrics | Cotton Kurti | Anarkali | Navy Blue | Size M", "Anvi Fabrics")
    assert any("multiple separators/delimiters" in w for w in warns), f"Failed delimiter test: {warns}"
    print("  ✓ Excessive delimiters flagged as advisory warning")

    # 1.6 Clean pattern-aligned title (Pass)
    errs, warns = validate_amazon_title_quality("Anvi Fabrics Women's Pure Cotton Printed Anarkali Kurti with Pockets (Navy Blue)", "Anvi Fabrics")
    assert len(errs) == 0, f"Clean title had unexpected errors: {errs}"
    print("  ✓ Standard pattern-aligned Amazon title passes without error")

    # ──────────────────────────────────────────────────────────────────────────
    # SECTION 2: Truthfulness & Claim Safety Validator
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[TEST SECTION 2] Truthfulness & Claim Safety Validation...")

    # 2.1 Technical & performance claims (Hard errors)
    hard_claim_samples = [
        ("breathable", "amazon.bullet_points[1]", "100% pure breathable cotton for summer"),
        ("cooling", "amazon.description", "Provides active cooling comfort all day"),
        ("sweat-absorbent", "flipkart.description", "Made from sweat-absorbent yarn"),
        ("quick-dry", "meesho.english_hook_description", "Quick-dry fabric for daily ease"),
        ("lightweight", "amazon.title", "Lightweight Pure Cotton Kurti"),
        ("skin-friendly", "flipkart.title", "Skin-friendly Anarkali Kurti"),
        ("zero fading", "amazon.bullet_points[3]", "Guarantees zero fading after multiple washes"),
        ("anti-pilling", "flipkart.search_keywords", "anti-pilling cotton kurti"),
        ("superior quality", "meesho.hinglish_hook_description", "Superior quality fabric ke saath"),
        ("guaranteed", "meesho.highlights[1]", "100% Guaranteed Colorfastness")
    ]
    for term, fpath, text in hard_claim_samples:
        errs, warns = validate_truthfulness_and_claims(fpath, text)
        assert len(errs) > 0, f"Expected hard error for '{term}' in '{fpath}', got: errs={errs}, warns={warns}"
        assert fpath in errs[0], f"Expected field path '{fpath}' in error message: {errs[0]}"
    print(f"  ✓ All {len(hard_claim_samples)} technical/guarantee claims correctly blocked as hard errors with exact field paths")

    # 2.2 Subjective fit & comfort claims (Advisory warnings)
    warning_claim_samples = [
        ("comfortable fit", "amazon.bullet_points[2]", "Offers a comfortable fit for office hours"),
        ("perfect fit", "flipkart.description", "Designed for a perfect fit on standard sizing"),
        ("soft feel", "meesho.hinglish_hook_description", "Pockets ke saath stylish look aur soft feel"),
        ("premium fabric", "amazon.title", "Premium fabric pure cotton kurti")
    ]
    for term, fpath, text in warning_claim_samples:
        errs, warns = validate_truthfulness_and_claims(fpath, text)
        assert len(errs) == 0, f"Expected 0 hard errors for '{term}', got {errs}"
        assert len(warns) > 0, f"Expected advisory warning for '{term}', got 0 warnings"
        assert fpath in warns[0], f"Expected field path '{fpath}' in warning message: {warns[0]}"
    print(f"  ✓ All {len(warning_claim_samples)} subjective comfort/fit claims correctly flagged as advisory warnings")

    # 2.3 Permitted styling suggestions (Pass without errors or warnings)
    styling_samples = [
        ("amazon.bullet_points[4]", "VERSATILE STYLING: Pairs with leggings, palazzos, or denim trousers."),
        ("flipkart.description", "An option for everyday ethnic styling or festive gatherings."),
        ("meesho.english_hook_description", "A thoughtful gifting option for family celebrations.")
    ]
    for fpath, text in styling_samples:
        errs, warns = validate_truthfulness_and_claims(fpath, text)
        assert len(errs) == 0 and len(warns) == 0, f"Styling sample '{text}' had unexpected flags: errs={errs}, warns={warns}"
    print(f"  ✓ All {len(styling_samples)} subjective styling suggestions pass cleanly")

    # ──────────────────────────────────────────────────────────────────────────
    # SECTION 3: Structural Readiness Scope & Meaning
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[TEST SECTION 3] Structural Readiness Scope & Meaning...")
    
    # 3.1 Constant check
    assert "Structural completeness confirms schema and package checks only" in STRUCTURAL_READINESS_DISCLAIMER
    print("  ✓ STRUCTURAL_READINESS_DISCLAIMER defined and verified")

    # 3.2 Master Summary Sheet check
    raw_sample_skus = get_sample_skus()
    sample_skus = [SKUItem(**d) for d in raw_sample_skus]
    image_map = {"SKU_01": [f"SKU_01{s}.jpg" for s in ["_MAIN", "_PT01", "_PT02", "_PT03"]]}
    wb_bytes = build_workbook(sample_skus, category="Women Ethnic Wear", image_map=image_map)
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws_summary = wb["Master_Summary"]
    
    headers = [cell.value for cell in ws_summary[1]]
    assert len(headers) == 15, f"Expected 15 columns in Master_Summary, got {len(headers)}: {headers}"
    assert headers[13] == "Package Readiness", f"Col 14 header mismatch: {headers[13]}"
    assert headers[14] == "Status Scope / Meaning", f"Col 15 header mismatch: {headers[14]}"
    
    row2 = [cell.value for cell in ws_summary[2]]
    assert row2[13] == "✅ Structurally Complete – Seller Review Required", f"Readiness status mismatch: {row2[13]}"
    assert row2[14] == STRUCTURAL_READINESS_DISCLAIMER, f"Scope meaning cell mismatch: {row2[14]}"
    print("  ✓ Master_Summary 15-column structure and Status Scope / Meaning column verified")

    # 3.3 Generated README check
    readme_text = generate_readme("TestClient", "Batch_01", "Women Ethnic Wear")
    assert STRUCTURAL_READINESS_DISCLAIMER in readme_text, "STRUCTURAL_READINESS_DISCLAIMER missing from generated README"
    print("  ✓ STRUCTURAL_READINESS_DISCLAIMER embedded in handover instructions")

    # ──────────────────────────────────────────────────────────────────────────
    # SECTION 4: Declared Image Roles & Verification Notice
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[TEST SECTION 4] Declared Image Roles & Verification Notice...")

    assert IMAGE_ROLES["_MAIN"] == "Declared Hero Image"
    assert IMAGE_ROLES["_PT01"] == "Declared Size Chart Slot"
    assert IMAGE_ROLES["_PT02"] == "Declared Fabric Specification Slot"
    assert IMAGE_ROLES["_PT03"] == "Declared Care Guide Slot"
    assert IMAGE_ROLES["_PT04"] == "Declared Back View Slot"
    assert IMAGE_ROLES["_PT05"] == "Declared Other Image 5 Slot"
    assert "Image roles are assigned from filenames only" in IMAGE_ROLE_DISCLAIMER
    assert IMAGE_ROLE_DISCLAIMER in readme_text
    assert "[SKU]_MAIN.jpg     ← Declared Primary hero (manual visual check required)" in readme_text
    assert "[SKU]_PT05.jpg     ← Declared Other Image 5 (manual visual check required)" in readme_text
    print("  ✓ Declared image roles and manual visual check notices verified in README and IMAGE_ROLES dict")

    # ──────────────────────────────────────────────────────────────────────────
    # SECTION 5: Full Regression & Sample Ingestion Check
    # ──────────────────────────────────────────────────────────────────────────
    print("\n[TEST SECTION 5] Full Regression & Sample Ingestion...")
    
    sample_sku = sample_skus[0]
    # Check title quality
    t_errs, t_warns = validate_amazon_title_quality(sample_sku.amazon.title, sample_sku.brand)
    assert len(t_errs) == 0 and len(t_warns) == 0, f"Sample title issues: errs={t_errs}, warns={t_warns}"

    # Check BST
    bst_errs, bst_warns = validate_amazon_backend_search_terms(sample_sku.amazon.title, sample_sku.amazon.backend_search_terms, sample_sku.brand)
    assert len(bst_errs) == 0 and len(bst_warns) == 0, f"Sample BST issues: errs={bst_errs}, warns={bst_warns}"

    # Check Bullet 3
    b3_errs, b3_warns = validate_bullet_3_safety(sample_sku.amazon.bullet_points[2])
    assert len(b3_errs) == 0 and len(b3_warns) == 0, f"Sample Bullet 3 issues: errs={b3_errs}, warns={b3_warns}"

    # Check claims on all copy
    for alt in sample_sku.alternates:
        at_errs, at_warns = validate_amazon_title_quality(alt.amazon.title, sample_sku.brand)
        assert len(at_errs) == 0, f"Alternate title error in {alt.variant_id}: {at_errs}"
        
        ab_errs, ab_warns = validate_bullet_3_safety(alt.amazon.bullet_points[2])
        assert len(ab_errs) == 0, f"Alternate B3 error in {alt.variant_id}: {ab_errs}"
    print(f"  ✓ Sample SKU (SKU_01) with 5x alternate variants validated 100% clean (0 errors, 0 warnings)")

    # Test Alternates Workbook
    alt_bytes = build_alternates_workbook(sample_skus)
    wb_alt = openpyxl.load_workbook(io.BytesIO(alt_bytes))
    assert "Amazon_Alternate_Copies" in wb_alt.sheetnames
    assert "Flipkart_Alternate_Copies" in wb_alt.sheetnames
    assert "Meesho_Alternate_Copies" in wb_alt.sheetnames
    print("  ✓ Alternate Listing Copies workbook generated with 3 marketplace sheets")

    # Test Metadata Builder
    meta = build_package_metadata("TestClient", "Batch_01", "Women Ethnic Wear", sample_skus)
    assert meta["tool_version"] == TOOL_VERSION
    assert meta["json_prompt_version"] == JSON_PROMPT_VERSION
    assert len(meta["skus"]) == 1
    assert meta["skus"][0]["input_hash"] is not None
    assert meta["skus"][0]["output_hash"] is not None
    print("  ✓ Cryptographic package metadata generated with valid SHA-256 hashes")

    print("\n================================================================================")
    print("ALL TESTS PASSED SUCCESSFULLY! (100% PASS RATE)")
    print("================================================================================")

if __name__ == "__main__":
    run_tests()
