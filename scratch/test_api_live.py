import os
import sys
import io
import json
import zipfile
import urllib.request
import urllib.error
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

BASE_URL = "http://127.0.0.1:8000"

def test_api():
    print("Testing FastAPI endpoints on", BASE_URL)
    
    # 1. Health / Home check
    req = urllib.request.Request(f"{BASE_URL}/")
    with urllib.request.urlopen(req) as resp:
        assert resp.status == 200
        print("  ✓ GET / returns 200 OK")

    # 2. Sample Generation Endpoint
    req = urllib.request.Request(f"{BASE_URL}/api/generate-sample", data=b"", method="POST")
    with urllib.request.urlopen(req) as resp:
        assert resp.status == 200
        assert resp.headers.get("content-type") == "application/zip"
        content = resp.read()
    
    # Inspect sample zip
    zf = zipfile.ZipFile(io.BytesIO(content))
    namelist = zf.namelist()
    print(f"  ✓ POST /api/generate-sample returned valid ZIP with {len(namelist)} entries")
    
    # Verify Master Summary in sample zip
    master_name = [n for n in namelist if n.endswith("_Master_Marketplace_Upload.xlsx")][0]
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(master_name)))
    ws = wb["Master_Summary"]
    assert len(ws[1]) == 15, f"Expected 15 cols, got {len(ws[1])}"
    assert ws[1][14].value == "Status Scope / Meaning"
    print("  ✓ Sample ZIP Master_Summary sheet has 15 columns with Status Scope / Meaning")

    # 3. Validation Endpoint - Clean sample payload
    from app import get_sample_skus, EXPECTED_SCHEMA_VERSION
    raw_sample = {
        "schema_version": EXPECTED_SCHEMA_VERSION,
        "batch_config": {
            "brand": "Anvi Fabrics",
            "category": "Women Ethnic Wear",
            "seller_config": {"amazon_quantity": 50, "gst_percent": 5, "hsn_code": "62114200"}
        },
        "skus": get_sample_skus()
    }
    payload = json.dumps({"json_text": json.dumps(raw_sample)}).encode("utf-8")
    req = urllib.request.Request(
        f"{BASE_URL}/api/validate-json",
        data=payload,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    with urllib.request.urlopen(req) as resp:
        assert resp.status == 200
        res = json.loads(resp.read().decode("utf-8"))
        assert res["valid"] is True
        print("  ✓ POST /api/validate-json validates sample payload with 100% pass (valid: true)")

    # 4. Validation Endpoint - Brand repetition error test
    bad_sample = json.loads(json.dumps(raw_sample))
    bad_sample["skus"][0]["amazon"]["title"] = "Anvi Fabrics Women Cotton Kurti by Anvi Fabrics"
    bad_payload = json.dumps({"json_text": json.dumps(bad_sample)}).encode("utf-8")
    req = urllib.request.Request(
        f"{BASE_URL}/api/validate-json",
        data=bad_payload,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        urllib.request.urlopen(req)
        assert False, "Expected 400 Bad Request"
    except urllib.error.HTTPError as e:
        assert e.code == 400
        res = json.loads(e.read().decode("utf-8"))
        assert res["valid"] is False
        assert any("brand name 'Anvi Fabrics' 2 times" in err for err in res["sku_results"][0]["errors"])
        print("  ✓ POST /api/validate-json correctly rejects brand duplication with 400 status")

    # 5. Validation Endpoint - Unverified breathable claim test
    bad_sample2 = json.loads(json.dumps(raw_sample))
    bad_sample2["skus"][0]["amazon"]["bullet_points"][0] = "100% Breathable Cotton: All-day airflow"
    bad_payload2 = json.dumps({"json_text": json.dumps(bad_sample2)}).encode("utf-8")
    req = urllib.request.Request(
        f"{BASE_URL}/api/validate-json",
        data=bad_payload2,
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    try:
        urllib.request.urlopen(req)
        assert False, "Expected 400 Bad Request"
    except urllib.error.HTTPError as e:
        assert e.code == 400
        res = json.loads(e.read().decode("utf-8"))
        assert res["valid"] is False
        assert any("breathable" in err for err in res["sku_results"][0]["errors"])
        print("  ✓ POST /api/validate-json correctly rejects unverified technical claim ('breathable') with 400 status")

    # 6. History Endpoint
    req = urllib.request.Request(f"{BASE_URL}/api/history")
    with urllib.request.urlopen(req) as resp:
        assert resp.status == 200
        data = json.loads(resp.read().decode("utf-8"))
        print(f"  ✓ GET /api/history returns 200 OK (found {len(data.get('packages', []))} packages)")

    print("\nALL API TESTS PASSED SUCCESSFULLY!")

if __name__ == "__main__":
    test_api()
